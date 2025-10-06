# interface.py
import os
import csv
import logging
from datetime import datetime
import sys
from pathlib import Path
from typing import Callable, TypedDict, Any, List, Dict, Optional, Sequence, cast, Tuple

from PyQt5.QtGui import (
    QFont,
    QPixmap,
    QRegularExpressionValidator,
    QStandardItemModel,
    QStandardItem,
)
from PyQt5.QtCore import (
    Qt,
    QRegularExpression,
    QObject,
    QSortFilterProxyModel,
    QTimer,
    # QProcess,
    QSettings,
    QCoreApplication,
    QEvent,
    QModelIndex,
    QTranslator,
)
from PyQt5.QtWidgets import (
    QMainWindow,
    QApplication,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QGridLayout,
    QGroupBox,
    QLabel,
    QLineEdit,
    QPushButton,
    QComboBox,
    QRadioButton,
    QButtonGroup,
    QMessageBox,
    QSizePolicy,
    QHeaderView,
    QTableView,
    QFrame,
    QAction,
    QActionGroup,
    QFileDialog,
    QTextEdit,
    QDialog,
    QDialogButtonBox,
)

import logic  # модуль расчётов


# Helper to resolve resource paths in both dev and PyInstaller onefile modes
def resource_path(*paths: str) -> Path:
    try:
        base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    except Exception:
        base = Path(__file__).resolve().parent
    return base.joinpath(*paths)


# Choose a writable directory for logs (e.g., %LOCALAPPDATA%\HeatSim\logs on Windows)
def writable_app_dir() -> Path:
    try:
        if sys.platform.startswith("win"):
            base_env = os.environ.get("LOCALAPPDATA") or os.path.join(
                str(Path.home()), "AppData", "Local"
            )
            base = Path(base_env)
        else:
            base = Path.home()
        return base / "HeatSim"
    except Exception:
        return Path.cwd() / "HeatSim"


LOG_DIR = writable_app_dir() / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = LOG_DIR / "app.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

# Feature flag: show legacy Language menu (kept for future use). Disabled per user request.
SHOW_LANGUAGE_MENU = False

try:
    import openpyxl  # type: ignore
except Exception:
    openpyxl = None
    # Workbook will be created via openpyxl.Workbook() when module is present

# ===================== БАЗА СВОЙСТВ КОМПОНЕНТОВ =====================
COMPONENT_DB: Dict[str, Tuple[float, float, float, float]] = {
    "Вода": (373.0, 4.2, 2.0, 2260.0),
    "Ртуть": (629.9, 0.14, 0.146, 294.0),
    "Этанол": (351.5, 2.44, 1.42, 846.0),
    "Азот": (77.4, 2.04, 1.04, 200.0),
    "Пропан": (231.0, 2.38, 1.67, 356.0),
    "Бутан": (272.7, 2.22, 1.67, 360.0),
    "Бензин": (388.0, 2.20, 1.70, 375.0),
    "Глицерин": (563.0, 2.43, 1.95, 924.0),
    "Фенол": (454.9, 2.10, 1.7, 654.0),
    "Водород": (20.2, 9.7, 14.3, 445.0),
    "Этиловый спирт": (351.5, 2.44, 1.42, 854.0),
    "Свинец": (2022.0, 0.15, 0.13, 871.0),
    "Аммиак": (239.8, 4.70, 2.09, 1370.0),
    "Медь": (2835.0, 0.62, 0.20, 4730.0),
    "Железо": (3135.0, 0.82, 0.45, 6770.0),
    "Алюминий": (2792.0, 1.18, 0.90, 10500.0),
    "Литий": (1615.0, 3.58, 3.58, 20200.0),
    "Графит": (4473.0, float("nan"), 0.71, 35500.0),
    "Диэтиловый эфир": (307.8, 2.19, 1.84, 412.0),
    "Бериллий": (2742.0, 1.82, 1.82, 12700.0),
    "Бор": (4200.0, 2.60, 1.02, 47000.0),
    "Сера": (718.0, 1.75, 0.71, 325.0),
    "Серная кислота": (610.0, 1.38, 1.40, 787.0),
    "Натрий": (1156.0, 1.25, 0.81, 8000.0),
    "Калий": (1032.0, 0.76, 0.75, 9560.0),
    "Хлор": (239.0, 0.48, 0.50, 287.0),
    "Йод": (457.0, 0.37, 0.17, 199.0),
    "Магний": (1363.0, 1.44, 1.02, 8571.0),
    "Кальций": (1757.0, 1.10, 0.65, 6970.0),
    "Цинк": (1180.0, 0.57, 0.52, 1700.0),
    "Олово": (2543.0, 0.30, 0.24, 2960.0),
    "Платина": (4100.0, 0.51, 0.13, 6000.0),
    "Никель": (3003.0, 0.75, 0.46, 6000.0),
    "Бензол": (353.25, 1.74, 1.13, 393.0),
    "Толуол": (383.75, 1.70, 1.13, 351.0),
    "Спирт": (351.52, 2.44, 1.43, 841.0),
    # --- Дополнения ---
    "Метанол": (337.85, 2.51, 1.95, 1100.0),
    "Изопропанол": (355.5, 2.68, 1.75, 667.0),
    "Ацетон": (329.45, 2.20, 1.58, 518.0),
    "Циклогексан": (353.87, 1.86, 1.12, 350.0),
    "Гексан (n-Hexane)": (341.88, 2.26, 1.67, 334.0),
    "Гептан (n-Heptane)": (371.58, 2.26, 1.66, 316.0),
    "Октан (n-Octane)": (398.83, 2.22, 1.64, 308.0),
    "Пентан (n-Pentane)": (309.21, 2.26, 1.69, 360.0),
    "Изобутан": (261.0, 2.28, 1.67, 366.0),
    "Пропилен (Пропен)": (225.5, 2.40, 1.68, 363.0),
    "Метан": (111.65, 3.50, 2.20, 510.0),
    "Этан": (184.55, 2.40, 1.74, 488.0),
    "Этен (Этилен)": (169.45, 2.35, 1.62, 430.0),
    "Кислород": (90.19, 1.70, 0.92, 213.0),
    "Аргон": (87.30, 1.30, 0.52, 161.0),
    "Этиленгликоль": (470.35, 2.42, 1.63, 800.0),
    "Пропиленгликоль": (460.35, 2.50, 1.60, 700.0),
    "R134a (1,1,1,2-ТФЭ)": (247.08, 1.42, 0.88, 216.0),
    "R32 (дифторметан)": (221.40, 1.77, 0.87, 238.0),
    "R22 (хлордифторметан)": (232.35, 1.31, 0.68, 233.0),
}

# Отображаемые английские названия компонентов для режима English (UI)
COMPONENT_NAME_EN: Dict[str, str] = {
    "Вода": "Water",
    "Ртуть": "Mercury",
    "Этанол": "Ethanol",
    "Азот": "Nitrogen",
    "Пропан": "Propane",
    "Бутан": "Butane",
    "Бензин": "Gasoline",
    "Глицерин": "Glycerin",
    "Фенол": "Phenol",
    "Водород": "Hydrogen",
    "Этиловый спирт": "Ethyl alcohol",
    "Свинец": "Lead",
    "Аммиак": "Ammonia",
    "Медь": "Copper",
    "Железо": "Iron",
    "Алюминий": "Aluminium",
    "Литий": "Lithium",
    "Графит": "Graphite",
    "Диэтиловый эфир": "Diethyl ether",
    "Бериллий": "Beryllium",
    "Бор": "Boron",
    "Сера": "Sulfur",
    "Серная кислота": "Sulfuric acid",
    "Натрий": "Sodium",
    "Калий": "Potassium",
    "Хлор": "Chlorine",
    "Йод": "Iodine",
    "Магний": "Magnesium",
    "Кальций": "Calcium",
    "Цинк": "Zinc",
    "Олово": "Tin",
    "Платина": "Platinum",
    "Никель": "Nickel",
    "Бензол": "Benzene",
    "Толуол": "Toluene",
    "Спирт": "Alcohol",
    "Метанол": "Methanol",
    "Изопропанол": "Isopropanol",
    "Ацетон": "Acetone",
    "Циклогексан": "Cyclohexane",
    "Гексан (n-Hexane)": "n-Hexane",
    "Гептан (n-Heptane)": "n-Heptane",
    "Октан (n-Octane)": "n-Octane",
    "Пентан (n-Pentane)": "n-Pentane",
    "Изобутан": "Isobutane",
    "Пропилен (Пропен)": "Propylene (Propene)",
    "Метан": "Methane",
    "Этан": "Ethane",
    "Этен (Этилен)": "Ethene (Ethylene)",
    "Кислород": "Oxygen",
    "Аргон": "Argon",
    "Этиленгликоль": "Ethylene glycol",
    "Пропиленгликоль": "Propylene glycol",
    "R134a (1,1,1,2-ТФЭ)": "R134a (1,1,1,2-Tetrafluoroethane)",
    "R32 (дифторметан)": "R32 (Difluoromethane)",
    "R22 (хлордифторметан)": "R22 (Chlorodifluoromethane)",
}
COMPONENT_NAME_RU_FROM_EN: Dict[str, str] = {v: k for k, v in COMPONENT_NAME_EN.items()}

# ---- Импорт/экспорт базы компонентов ----
DATA_DIR = Path(os.path.dirname(os.path.abspath(__file__))) / "data"


def _parse_float_cell(val: Optional[str]) -> Optional[float]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return float(s.replace(",", "."))
    except Exception:
        return None


def load_component_db_from_xlsx(
    path: os.PathLike[str] | str, merge: bool = True
) -> Dict[str, int]:
    """Load component properties from an Excel .xlsx file.

    Expected header names are the same as CSV loader (Russian/English variants):
    имя|name, Tb_K|Tb, Cf_kJ_per_kgK|C_f|Cf, Cp_kJ_per_kgK|C_p|Cp, rf_kJ_per_kg|r_f|rf
    Only numeric columns are used; any extra columns (e.g., source_url) are ignored.
    """
    if openpyxl is None:
        raise RuntimeError("Для импорта из Excel требуется пакет openpyxl.")
    stats = {"updated": 0, "added": 0, "skipped": 0}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        # use sheet named 'components' if exists, else first sheet
        ws = None
        if "components" in wb.sheetnames:
            ws = wb["components"]
        else:
            ws = wb[wb.sheetnames[0]]
        rows = list(ws.values)
        if not rows:
            return stats
        # header row
        header = [str(h or "").strip() for h in rows[0]]

        def hfind(*options: str) -> Optional[int]:
            low = [h.lower() for h in header]
            for o in options:
                if o.lower() in low:
                    return low.index(o.lower())
            return None

        i_name = hfind("имя", "name")
        i_tb = hfind("tb_k", "tb")
        i_cf = hfind("cf_kj_per_kgk", "c_f", "cf")
        i_cp = hfind("cp_kj_per_kgk", "c_p", "cp")
        i_rf = hfind("rf_kj_per_kg", "r_f", "rf")
        if i_name is None:
            raise ValueError("В Excel отсутствует столбец 'имя'/'name'.")
        for r in rows[1:]:
            try:
                name = str(r[i_name] or "").strip() if i_name < len(r) else ""
                if not name:
                    stats["skipped"] += 1
                    continue

                def getf(idx: Optional[int]) -> Optional[float]:
                    if idx is None or idx >= len(r):
                        return None
                    v = r[idx]
                    if v is None or v == "":
                        return None
                    try:
                        return float(str(v).replace(",", "."))
                    except Exception:
                        return None

                tb = getf(i_tb)
                cf = getf(i_cf)
                cp = getf(i_cp)
                rf = getf(i_rf)
                has_any = any(v is not None for v in (tb, cf, cp, rf))
                all_present = all(v is not None for v in (tb, cf, cp, rf))
                if not has_any:
                    stats["skipped"] += 1
                    continue
                if name in COMPONENT_DB:
                    old_tb, old_cf, old_cp, old_rf = COMPONENT_DB[name]
                    if merge:
                        new_val = (
                            tb if tb is not None else old_tb,
                            cf if cf is not None else old_cf,
                            cp if cp is not None else old_cp,
                            rf if rf is not None else old_rf,
                        )
                        if new_val != COMPONENT_DB[name]:
                            COMPONENT_DB[name] = new_val
                            stats["updated"] += 1
                        else:
                            stats["skipped"] += 1
                    else:
                        if all_present:
                            COMPONENT_DB[name] = (tb, cf, cp, rf)  # type: ignore[arg-type]
                            stats["updated"] += 1
                        else:
                            stats["skipped"] += 1
                else:
                    if all_present:
                        COMPONENT_DB[name] = (tb, cf, cp, rf)  # type: ignore[arg-type]
                        stats["added"] += 1
                    else:
                        stats["skipped"] += 1
            except Exception:
                stats["skipped"] += 1
                continue
    except Exception as e:
        logger.exception("Ошибка импорта базы компонентов из Excel: %s", e)
        raise
    return stats


def load_component_db_from_csv(
    path: os.PathLike[str] | str, merge: bool = True
) -> Dict[str, int]:
    """Load component properties from CSV.

    Expected columns (case-insensitive, Russian/English supported):
    - имя | name
    - Tb_K | Tb
    - Cf_kJ_per_kgK | C_f
    - Cp_kJ_per_kgK | C_p
    - rf_kJ_per_kg | r_f

    Behavior:
    - merge=True: for existing components update only provided fields; for new components add only if all 4 fields are present.
    - merge=False: replace existing entries entirely when all 4 fields present.

    Returns stats dict: {"updated": x, "added": y, "skipped": z}.
    """
    stats = {"updated": 0, "added": 0, "skipped": 0}
    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except Exception:
                # часто используется ';' в проекте
                dialect = csv.excel
                dialect.delimiter = ";"  # type: ignore[attr-defined]
            rdr = csv.DictReader(f, dialect=dialect)
            # нормализуем заголовки
            headers = {(h or "").strip().lower(): h for h in (rdr.fieldnames or [])}

            def hkey(*options: str) -> Optional[str]:
                for o in options:
                    if o.lower() in headers:
                        return headers[o.lower()]
                return None

            col_name = hkey("имя", "name")
            col_tb = hkey("tb_k", "tb")
            col_cf = hkey("cf_kj_per_kgk", "c_f", "cf")
            col_cp = hkey("cp_kj_per_kgk", "c_p", "cp")
            col_rf = hkey("rf_kj_per_kg", "r_f", "rf")

            if not col_name:
                raise ValueError("В CSV отсутствует столбец 'имя'/'name'.")

            for row in rdr:
                name = (row.get(col_name) or "").strip()
                if not name:
                    stats["skipped"] += 1
                    continue
                tb = _parse_float_cell(row.get(col_tb)) if col_tb else None
                cf = _parse_float_cell(row.get(col_cf)) if col_cf else None
                cp = _parse_float_cell(row.get(col_cp)) if col_cp else None
                rf = _parse_float_cell(row.get(col_rf)) if col_rf else None

                has_any = any(v is not None for v in (tb, cf, cp, rf))
                all_present = all(v is not None for v in (tb, cf, cp, rf))
                if not has_any:
                    stats["skipped"] += 1
                    continue

                if name in COMPONENT_DB:
                    old_tb, old_cf, old_cp, old_rf = COMPONENT_DB[name]
                    if merge:
                        new_val = (
                            tb if tb is not None else old_tb,
                            cf if cf is not None else old_cf,
                            cp if cp is not None else old_cp,
                            rf if rf is not None else old_rf,
                        )
                        if new_val != COMPONENT_DB[name]:
                            COMPONENT_DB[name] = new_val
                            stats["updated"] += 1
                        else:
                            stats["skipped"] += 1
                    else:
                        if all_present:
                            COMPONENT_DB[name] = (tb, cf, cp, rf)  # type: ignore[arg-type]
                            stats["updated"] += 1
                        else:
                            stats["skipped"] += 1
                else:
                    if all_present:
                        COMPONENT_DB[name] = (tb, cf, cp, rf)  # type: ignore[arg-type]
                        stats["added"] += 1
                    else:
                        stats["skipped"] += 1
    except FileNotFoundError:
        raise
    except Exception as e:
        logger.exception("Ошибка импорта базы компонентов: %s", e)
        raise
    return stats


# CSV-экспорт базы компонентов удалён по требованию; используйте Excel-экспорт.


def export_component_db_to_xlsx(path: os.PathLike[str] | str) -> None:
    """Export current COMPONENT_DB to an Excel .xlsx workbook.

    Sheet name: components
    Columns: имя, Tb_K, Cf_kJ_per_kgK, Cp_kJ_per_kgK, rf_kJ_per_kg, source_url
    """
    if openpyxl is None:
        raise RuntimeError("Для экспорта в Excel требуется пакет openpyxl.")

    # Build URL map from existing CSV mapping if present
    url_map: Dict[str, str] = {}
    nist_csv = DATA_DIR / "components_nist_results.csv"
    try:
        if nist_csv.exists():
            with open(nist_csv, "r", encoding="utf-8-sig", newline="") as f:
                rdr = csv.DictReader(f)
                name_col = None
                url_col = None
                if rdr.fieldnames:
                    fl = [h.strip().lower() for h in rdr.fieldnames]
                    for i, h in enumerate(fl):
                        if h in ("имя", "name"):
                            name_col = rdr.fieldnames[i]
                        if h in ("source_url", "url"):
                            url_col = rdr.fieldnames[i]
                for row in rdr:
                    nm = (row.get(name_col or "") or "").strip()
                    url = (row.get(url_col or "") or "").strip()
                    if nm and url:
                        url_map[nm] = url
    except Exception:
        pass

    assert openpyxl is not None
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.title = "components"
    ws.append(
        ["имя", "Tb_K", "Cf_kJ_per_kgK", "Cp_kJ_per_kgK", "rf_kJ_per_kg", "source_url"]
    )
    for name in sorted(COMPONENT_DB.keys()):
        tb, cf, cp, rf = COMPONENT_DB[name]
        url = url_map.get(name, "")
        ws.append([name, tb, cf, cp, rf, url])
    wb.save(path)


def _auto_load_components_db() -> Optional[Dict[str, int]]:
    """Auto-load DB from known CSV if present. Returns stats or None."""
    candidates = [
        DATA_DIR / "components.xlsx",
        DATA_DIR / "components.csv",
        DATA_DIR / "components_nist_results.csv",
    ]
    for p in candidates:
        try:
            if p.exists():
                if p.suffix.lower() == ".xlsx":
                    return load_component_db_from_xlsx(p, merge=True)
                else:
                    return load_component_db_from_csv(p, merge=True)
        except Exception:
            # не критично для старта приложения
            return None
    return None


# Выполним авто-загрузку на старте (не критично при ошибке)
try:
    _ = _auto_load_components_db()
except Exception:
    pass


# ===================== УТИЛИТЫ =====================
def num_edit(read_only: bool = False, fixed_width: int = 150) -> QLineEdit:
    e = QLineEdit()
    e.setAlignment(Qt.AlignRight)
    # initial enabled state will be controlled via set_enabled helper
    e.setEnabled(not read_only)
    e.setPlaceholderText("0.0")
    e.setFixedWidth(fixed_width)
    if not e.isEnabled():
        e.setStyleSheet("background:#f3f3f3;")
    rx = QRegularExpression(r"^$|^[0-9]{1,10}([.,][0-9]{0,5})?$")
    e.setValidator(QRegularExpressionValidator(rx))

    def fix_number():
        t = e.text().strip()
        if not t:
            return
        if t.endswith(",") or t.endswith("."):
            t += "00"
        sep = max(t.rfind(","), t.rfind("."))
        if sep != -1:
            i, f = t[:sep], t[sep + 1 :]
            t = i[:10] + t[sep] + (f or "00")[:5]
        else:
            t = t[:10]
        e.blockSignals(True)
        e.setText(t)
        e.blockSignals(False)

    e.editingFinished.connect(fix_number)
    return e


def to_float(text: str) -> float:
    try:
        return float(text.replace(",", "."))
    except Exception:
        return 0.0


def format_num(value: float, fmt: str = ".6g") -> str:
    try:
        v = float(value)
        # special-case exact zero for clearer display
        if abs(v) < 1e-12:
            return "0.0"
        return f"{v:{fmt}}"
    except Exception:
        return "0.0"


def set_read_only(le: QLineEdit, ro: bool) -> None:
    # kept for backward-compat but delegate to new enabled semantics
    set_enabled(le, not ro)


def set_enabled(le: QLineEdit, enabled: bool) -> None:
    le.setEnabled(enabled)
    if not enabled:
        # disabled fields: gray background
        le.setStyleSheet("background:#f3f3f3;")
    else:
        # enabled fields: clear style (caller may set manual highlight)
        le.setStyleSheet("")
    # keep associated lock button in sync without triggering its signals
    btn = getattr(le, "_lock_btn", None)
    if btn is not None:
        # update text only; button click will set enabled state explicitly
        btn.setText("🔒" if not enabled else "🔓")


def lock_button_for(line_edit: QLineEdit) -> QPushButton:
    btn = QPushButton()
    btn.setFixedSize(22, 22)
    btn.setToolTip(
        QCoreApplication.translate("UI", "Заблокировать/разблокировать поле")
    )

    def on_click():
        # if the field is enabled -> lock it; otherwise unlock
        if line_edit.isEnabled():
            # lock
            set_enabled(line_edit, False)
            setattr(line_edit, "_just_unlocked", False)
            # remove any temporary handler if present
            h = getattr(line_edit, "_just_unlocked_handler", None)
            if h is not None:
                try:
                    line_edit.textEdited.disconnect(h)
                except Exception:
                    pass
                try:
                    delattr(line_edit, "_just_unlocked_handler")
                except Exception:
                    pass
        else:
            # unlock — prepare flags so that an immediate editingFinished (without user typing)
            # won't auto-disable, but a real typed edit followed by editingFinished will.
            set_enabled(line_edit, True)
            try:
                # waiting flag indicates we recently unlocked and expect possible typing
                setattr(line_edit, "_just_unlocked_waiting", True)
                # clear typed flag
                if hasattr(line_edit, "_just_unlocked_typed"):
                    delattr(line_edit, "_just_unlocked_typed")
                # remember initial text to detect changes even if textEdited didn't fire
                try:
                    setattr(line_edit, "_unlock_initial_text", line_edit.text())
                except Exception:
                    setattr(line_edit, "_unlock_initial_text", None)
            except Exception:
                pass

            def _on_text_edited(_text: str) -> None:
                # mark that the user actually typed
                try:
                    setattr(line_edit, "_just_unlocked_typed", True)
                finally:
                    try:
                        line_edit.textEdited.disconnect(_on_text_edited)
                    except Exception:
                        pass
                    try:
                        if hasattr(line_edit, "_just_unlocked_handler"):
                            delattr(line_edit, "_just_unlocked_handler")
                    except Exception:
                        pass

            # store handler reference for cleanup and connect
            try:
                setattr(line_edit, "_just_unlocked_handler", _on_text_edited)
                line_edit.textEdited.connect(_on_text_edited)
            except Exception:
                try:
                    if hasattr(line_edit, "_just_unlocked_handler"):
                        delattr(line_edit, "_just_unlocked_handler")
                except Exception:
                    pass

    btn.clicked.connect(on_click)
    # initial text reflects current state
    btn.setText("🔒" if not line_edit.isEnabled() else "🔓")
    # attach for external sync
    setattr(line_edit, "_lock_btn", btn)
    return btn


def auto_disable_handler(line_edit: QLineEdit) -> Callable[[], None]:
    def _handler() -> None:
        # if we just unlocked for editing, only skip auto-disable when no typing occurred
        if getattr(line_edit, "_just_unlocked_waiting", False):
            # decide if text actually changed since unlock
            try:
                initial_txt = getattr(line_edit, "_unlock_initial_text", None)
                current_txt = line_edit.text()
                changed = (initial_txt is not None) and (current_txt != initial_txt)
            except Exception:
                changed = False
            # if user typed or text changed, proceed to disable and clear flags
            if getattr(line_edit, "_just_unlocked_typed", False) or changed:
                try:
                    delattr(line_edit, "_just_unlocked_typed")
                except Exception:
                    pass
                try:
                    delattr(line_edit, "_just_unlocked_waiting")
                except Exception:
                    pass
                try:
                    if hasattr(line_edit, "_unlock_initial_text"):
                        delattr(line_edit, "_unlock_initial_text")
                except Exception:
                    pass
                # allow auto-disable to proceed
            else:
                # user didn't type yet — skip disabling for now
                return
        set_enabled(line_edit, False)

    return _handler


# ===================== TYPE DEFINITIONS =====================
class FlowData(TypedDict):
    t_in: float
    t_out: float
    m: float
    p: float


class MixRow(TypedDict):
    name: str
    share: float
    tb: float
    cf: float
    cp: float
    rf: float


class CalcResult(TypedDict, total=False):
    q: float
    t_out_plus: float
    sigma: float
    k: float


# ===================== ПАНЕЛЬ ПОТОКОВ =====================
class FlowPanel:
    def __init__(self, title: str, sign: str):
        self._title_ru = title
        self._sign = sign
        self.box = QGroupBox(title)
        grid = QGridLayout(self.box)
        self.t_in = num_edit()
        self.t_out = num_edit()
        self.m = num_edit()
        self.p = num_edit()

        # per-field lock buttons for some inputs (auto-disable after editing)
        self.t_in_lock = lock_button_for(self.t_in)
        self.t_out_lock = lock_button_for(self.t_out)
        self.m_lock = lock_button_for(self.m)
        self.p_lock = lock_button_for(self.p)

        row = 0
        self.lbl_t_in = QLabel(
            QCoreApplication.translate(
                "FlowPanel",
                "Температура на входе ({part}), T{sub} [ K ]",
            ).format(part=title.lower(), sub=f"<sub>{sign}</sub><sup>in</sup>")
        )
        grid.addWidget(self.lbl_t_in, row, 0)
        h0 = QHBoxLayout()
        h0.setContentsMargins(0, 0, 0, 0)
        h0.addWidget(self.t_in)
        h0.addWidget(self.t_in_lock)
        grid.addLayout(h0, row, 1)

        row += 1
        self.lbl_t_out = QLabel(
            QCoreApplication.translate(
                "FlowPanel",
                "Температура на выходе ({part}), T{sub} [ K ]",
            ).format(part=title.lower(), sub=f"<sub>{sign}</sub><sup>out</sup>")
        )
        grid.addWidget(self.lbl_t_out, row, 0)
        h1 = QHBoxLayout()
        h1.setContentsMargins(0, 0, 0, 0)
        h1.addWidget(self.t_out)
        h1.addWidget(self.t_out_lock)
        grid.addLayout(h1, row, 1)

        row += 1
        self.lbl_m = QLabel(
            QCoreApplication.translate(
                "FlowPanel",
                "Расход потока ({part}), g{sub} [ кг/сек ]",
            ).format(part=title.lower(), sub=f"<sub>{sign}</sub>")
        )
        grid.addWidget(self.lbl_m, row, 0)
        h2 = QHBoxLayout()
        h2.setContentsMargins(0, 0, 0, 0)
        h2.addWidget(self.m)
        h2.addWidget(self.m_lock)
        grid.addLayout(h2, row, 1)

        row += 1
        self.lbl_p = QLabel(
            QCoreApplication.translate(
                "FlowPanel",
                "Давление ({part}), P{sub} [ кг/м² ]",
            ).format(part=title.lower(), sub=f"<sub>{sign}</sub>")
        )
        grid.addWidget(self.lbl_p, row, 0)
        h3 = QHBoxLayout()
        h3.setContentsMargins(0, 0, 0, 0)
        h3.addWidget(self.p)
        h3.addWidget(self.p_lock)
        grid.addLayout(h3, row, 1)

        # Расширяемая по горизонтали панель, фиксируем только высоту
        self.box.setFixedHeight(180)
        self.box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(8)

        # auto-disable these fields after editingFinished (user can re-enable with lock button)
        try:
            self.t_in.editingFinished.connect(auto_disable_handler(self.t_in))
            self.t_out.editingFinished.connect(auto_disable_handler(self.t_out))
            self.m.editingFinished.connect(auto_disable_handler(self.m))
            self.p.editingFinished.connect(auto_disable_handler(self.p))
        except Exception:
            pass

    def widget(self) -> QGroupBox:
        return self.box

    def to_dict(self) -> FlowData:
        return FlowData(
            {
                "t_in": to_float(self.t_in.text()),
                "t_out": to_float(self.t_out.text()),
                "m": to_float(self.m.text()),
                "p": to_float(self.p.text()),
            }
        )

    def _localized_titles(self, lang: str) -> tuple[str, str]:
        lang = (lang or "ru").lower()
        if lang.startswith("en"):
            part = "hot flow" if self._sign == "+" else "cold flow"
            title = "Hot flow" if self._sign == "+" else "Cold flow"
            return title, part
        return self._title_ru, self._title_ru.lower()

    def retranslate_panel(self, lang: str) -> None:
        try:
            title, part = self._localized_titles(lang)
            self.box.setTitle(title)
            if str(lang or "").lower().startswith("en"):
                self.lbl_t_in.setText(
                    f"Inlet temperature ({part}), T<sub>{self._sign}</sub><sup>in</sup> [ K ]"
                )
                self.lbl_t_out.setText(
                    f"Outlet temperature ({part}), T<sub>{self._sign}</sub><sup>out</sup> [ K ]"
                )
                self.lbl_m.setText(
                    f"Flow rate ({part}), g<sub>{self._sign}</sub> [ kg/s ]"
                )
                self.lbl_p.setText(
                    f"Pressure ({part}), P<sub>{self._sign}</sub> [ kg/m² ]"
                )
            else:
                self.lbl_t_in.setText(
                    QCoreApplication.translate(
                        "FlowPanel",
                        "Температура на входе ({part}), T{sub} [ K ]",
                    ).format(part=part, sub=f"<sub>{self._sign}</sub><sup>in</sup>")
                )
                self.lbl_t_out.setText(
                    QCoreApplication.translate(
                        "FlowPanel",
                        "Температура на выходе ({part}), T{sub} [ K ]",
                    ).format(part=part, sub=f"<sub>{self._sign}</sub><sup>out</sup>")
                )
                self.lbl_m.setText(
                    QCoreApplication.translate(
                        "FlowPanel",
                        "Расход потока ({part}), g{sub} [ кг/сек ]",
                    ).format(part=part, sub=f"<sub>{self._sign}</sub>")
                )
                self.lbl_p.setText(
                    QCoreApplication.translate(
                        "FlowPanel",
                        "Давление ({part}), P{sub} [ кг/м² ]",
                    ).format(part=part, sub=f"<sub>{self._sign}</sub>")
                )
        except Exception:
            pass


# ===================== DELETE FILTER =====================
class KeyDeleteFilter(QObject):
    def __init__(self, callback: Callable[[], None]):
        super().__init__()
        self.callback: Callable[[], None] = callback

    def eventFilter(self, obj: QObject, event: QEvent) -> bool:  # type: ignore[override]
        if (
            event.type() == QEvent.KeyPress
            and getattr(event, "key", lambda: None)() == Qt.Key_Delete
        ):
            self.callback()
            return True
        return super().eventFilter(obj, event)


# ===================== ФИЛЬТР АВТО-БЛОКИРОВКИ НА ВЫХОДЕ ИЗ ФОКУСА =====================
class AutoLockRecalcFilter(QObject):
    """Фильтр, который при уходе фокуса:
    - при необходимости авто-блокирует поле (как auto_disable_handler)
    - если был явный расчёт и значение изменили, показывает кнопку «Перерасчёт»
    """

    def __init__(
        self,
        owner: QMainWindow,
        line_edit: QLineEdit,
        on_any_input_changed: Callable[[], None],
    ) -> None:
        super().__init__(owner)
        self._owner = owner
        self._le = line_edit
        self._on_changed = on_any_input_changed

    def eventFilter(self, obj: QObject, event: QEvent) -> bool:  # type: ignore[override]
        try:
            if obj is self._le and event.type() == QEvent.FocusOut:
                # Определим, нужно ли авто-блокировать
                do_lock = True
                if getattr(self._le, "_just_unlocked_waiting", False):
                    try:
                        initial_txt = getattr(self._le, "_unlock_initial_text", None)
                        current_txt = self._le.text()
                        changed = (initial_txt is not None) and (
                            current_txt != initial_txt
                        )
                    except Exception:
                        changed = False
                    # Если пользователь не вводил и текст не изменился — не блокируем
                    if (
                        not getattr(self._le, "_just_unlocked_typed", False)
                        and not changed
                    ):
                        do_lock = False
                    # Сбрасываем флаги, если будем блокировать
                    if do_lock:
                        for attr in (
                            "_just_unlocked_typed",
                            "_just_unlocked_waiting",
                            "_unlock_initial_text",
                        ):
                            try:
                                if hasattr(self._le, attr):
                                    delattr(self._le, attr)
                            except Exception:
                                pass
                if do_lock and self._le.isEnabled():
                    # Авто-блокировка поля
                    set_enabled(self._le, False)
                    # Если уже был явный расчёт — показать кнопку «Перерасчёт»
                    try:
                        if getattr(self._owner, "_explicit_calc_done", False):
                            self._on_changed()
                    except Exception:
                        pass
        except Exception:
            pass
        return super().eventFilter(obj, event)


# ===================== МОДЕЛЬ СМЕСИ =====================
class MixModel(QStandardItemModel):
    COL_NAME, COL_SHARE, COL_TB, COL_CF, COL_CP, COL_RF = range(6)
    HEADERS = [
        "Компонент",
        "Доля",
        "Tb, K",
        "C_f, кДж/кг·K",
        "C_p, кДж/кг·K",
        "r_f, кДж/кг",
    ]
    SORT_ROLE = Qt.UserRole + 1

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(0, 6, parent)
        for i, h in enumerate(self.HEADERS):
            self.setHeaderData(
                i,
                Qt.Horizontal,
                QCoreApplication.translate("MixModel", h),
                role=Qt.DisplayRole,
            )

    def retranslate_headers(self) -> None:
        """Переустанавливает заголовки столбцов и даёт явные EN подписи без .qm."""
        # Определяем язык аналогично MixPanel._is_lang_en
        en = False
        try:
            app_inst = QApplication.instance()
            active_lang = (
                str(getattr(app_inst, "_app_translator_lang", "") or "").lower()
                if app_inst is not None
                else ""
            )
            desired = str(QSettings().value("ui/language", "ru") or "ru").lower()
            en = desired == "en" or active_lang.startswith("en")
        except Exception:
            en = False
        headers_en = [
            "Component",
            "Share",
            "Tb, K",
            "C_f, kJ/kg·K",
            "C_p, kJ/kg·K",
            "r_f, kJ/kg",
        ]
        for i, h in enumerate(self.HEADERS):
            try:
                text = (
                    headers_en[i] if en else QCoreApplication.translate("MixModel", h)
                )
                self.setHeaderData(i, Qt.Horizontal, text, role=Qt.DisplayRole)
            except Exception:
                pass

    def _num_item(self, value: float) -> QStandardItem:
        it = QStandardItem(f"{value:.6g}")
        it.setEditable(False)
        it.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        it.setData(value, self.SORT_ROLE)
        return it

    def _set_num(self, row: int, col: int, value: float) -> None:
        idx = self.index(row, col)
        self.setData(idx, f"{value:.6g}", Qt.DisplayRole)
        self.setData(idx, value, self.SORT_ROLE)

    def add_or_update(
        self, name: str, share: float, tb: float, cf: float, cp: float, rf: float
    ) -> int:
        row = self._row_by_name(name)
        if row >= 0:
            idx_share = self.index(row, self.COL_SHARE)
            cur_share = float(self.data(idx_share, Qt.DisplayRole).replace(",", "."))
            new_share = cur_share + share
            self.setData(idx_share, f"{new_share:.6g}", Qt.DisplayRole)
            self.setData(idx_share, new_share, self.SORT_ROLE)
            self._set_num(row, self.COL_TB, tb)
            self._set_num(row, self.COL_CF, cf)
            self._set_num(row, self.COL_CP, cp)
            self._set_num(row, self.COL_RF, rf)
            return row
        r = self.rowCount()
        self.insertRow(r)
        name_item = QStandardItem(name)
        name_item.setEditable(False)
        name_item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        self.setItem(r, self.COL_NAME, name_item)
        self.setItem(r, self.COL_SHARE, self._num_item(share))
        self.setItem(r, self.COL_TB, self._num_item(tb))
        self.setItem(r, self.COL_CF, self._num_item(cf))
        self.setItem(r, self.COL_CP, self._num_item(cp))
        self.setItem(r, self.COL_RF, self._num_item(rf))
        return r

    def _row_by_name(self, name: str) -> int:
        for r in range(self.rowCount()):
            if self.data(self.index(r, self.COL_NAME), Qt.DisplayRole) == name:
                return r
        return -1

    def remove_rows(self, rows: List[int]) -> None:
        for r in sorted(rows, reverse=True):
            self.removeRow(r)

    def rows_as_dicts(self) -> List[MixRow]:
        out: List[MixRow] = []
        for r in range(self.rowCount()):

            def v(c: int):
                txt = self.data(self.index(r, c), Qt.DisplayRole) or "0"
                return float(txt.replace(",", ".")) if c != self.COL_NAME else txt

            out.append(
                cast(
                    MixRow,
                    {
                        "name": v(self.COL_NAME),
                        "share": v(self.COL_SHARE),
                        "tb": v(self.COL_TB),
                        "cf": v(self.COL_CF),
                        "cp": v(self.COL_CP),
                        "rf": v(self.COL_RF),
                    },
                )
            )
        return out


# ===================== ПАНЕЛЬ СМЕСИ =====================
class MixPanel:
    def __init__(self, title: str, is_hot: bool, export_path: str):
        self.is_hot = is_hot
        # сохранить русское название части потока для локализации заголовка
        self._title_ru = title
        self.export_path = export_path
        self.box = QGroupBox(
            QCoreApplication.translate("MixPanel", "Смесь компонентов {part}").format(
                part=title.lower()
            )
        )
        self.box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        v = QVBoxLayout(self.box)

        # верхняя линия управления
        top = QHBoxLayout()
        top.setContentsMargins(0, 0, 0, 0)
        top.setSpacing(6)
        self.comp = QComboBox()
        self._populate_component_combo()
        self.comp.setFixedWidth(300)
        self.comp.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.share = num_edit(fixed_width=100)
        self.share.editingFinished.connect(self.validate_share_max1)
        self.sum_field = num_edit(read_only=True, fixed_width=100)
        self.sum_field.setText("0.0")
        self.add_btn = QPushButton(self.box.tr("Добавить"))
        top.addWidget(self.comp)
        top.addStretch(1)
        self.lbl_share = QLabel(self.box.tr("Доля"))
        top.addWidget(self.lbl_share)
        top.addWidget(self.share)
        top.addSpacing(8)
        # Перестановка: сначала кнопка Добавить, затем поле суммы
        top.addWidget(self.add_btn)
        top.addSpacing(8)
        self.lbl_sum = QLabel(self.box.tr("Сумма"))
        top.addWidget(self.lbl_sum)
        top.addWidget(self.sum_field)
        v.addLayout(top)

        # источник параметров
        src = QHBoxLayout()
        src.setContentsMargins(0, 0, 0, 0)
        src.setSpacing(8)
        self.rb_group = QButtonGroup(self.box)
        self.rb_db = QRadioButton(
            self.box.tr("Взять параметры из справочника NIST Chemistry WebBook")
        )
        self.rb_manual = QRadioButton(self.box.tr("Ввести параметры вручную"))
        self.rb_group.addButton(self.rb_db, 0)
        self.rb_group.addButton(self.rb_manual, 1)
        self.rb_db.setChecked(True)
        src.addWidget(self.rb_db)
        src.addWidget(self.rb_manual)
        src.addStretch(1)
        v.addLayout(src)

        # параметры
        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(6)
        self.tb = num_edit(read_only=True)
        self.cf = num_edit(read_only=True)
        self.cp = num_edit(read_only=True)
        self.rf = num_edit(read_only=True)
        self.lbl_tb = QLabel(self.box.tr("Температура кипения, Tb  [ K ]"))
        grid.addWidget(self.lbl_tb, 0, 0)
        grid.addWidget(self.tb, 0, 1)
        self.lbl_cf = QLabel(
            self.box.tr("Удельная теплоёмкость жидкости, C_f  [ кДж/кг·K ]")
        )
        grid.addWidget(
            self.lbl_cf,
            1,
            0,
        )
        grid.addWidget(self.cf, 1, 1)
        self.lbl_cp = QLabel(
            self.box.tr("Удельная теплоёмкость пара, C_p  [ кДж/кг·K ]")
        )
        grid.addWidget(self.lbl_cp, 2, 0)
        grid.addWidget(self.cp, 2, 1)
        self.lbl_rf = QLabel(
            self.box.tr("Скрытая теплота фазового перехода, r_f  [ кДж/кг ]")
        )
        grid.addWidget(
            self.lbl_rf,
            3,
            0,
        )
        grid.addWidget(self.rf, 3, 1)
        v.addLayout(grid)

        # таблица
        self.model = MixModel()
        self.proxy = QSortFilterProxyModel()
        self.proxy.setSourceModel(self.model)
        self.proxy.setSortRole(MixModel.SORT_ROLE)
        self.proxy.setDynamicSortFilter(True)
        self.view = QTableView()
        self.view.setModel(self.proxy)
        self.view.setSortingEnabled(False)
        self.view.horizontalHeader().setVisible(True)
        self.view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.view.verticalHeader().setVisible(False)
        self.view.setSelectionBehavior(QTableView.SelectRows)
        self.view.setSelectionMode(QTableView.ExtendedSelection)
        v.addWidget(self.view)

        # шрифты
        header_font = QFont("Consolas", 9, QFont.Bold)
        self.view.horizontalHeader().setFont(header_font)
        try:
            hh = self.view.horizontalHeader()
            hh_font = hh.font()
            hh_font.setBold(True)
            hh.setFont(hh_font)
        except Exception:
            pass
        table_font = QFont("Consolas", 9)
        self.view.setFont(table_font)
        self.view.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        try:
            self.view.setStyleSheet("QHeaderView::section { font-weight: 700; }")
        except Exception:
            pass

        # Delete + двойной клик
        self._del_filter = KeyDeleteFilter(self.delete_selected_rows)
        self.view.installEventFilter(self._del_filter)
        self.view.doubleClicked.connect(self.on_double_click)

        # автоэкспорт и пересчёт
        self.model.dataChanged.connect(self._on_model_changed)
        self.model.rowsInserted.connect(self._on_model_changed)
        self.model.rowsRemoved.connect(self._on_model_changed)

        # сигналы
        self.add_btn.clicked.connect(self.on_add)
        self.rb_db.toggled.connect(self.on_mode_change)
        self.comp.currentIndexChanged.connect(lambda _ix: self.fill_from_db(self._current_component_key()))  # type: ignore
        self.fill_from_db(self._current_component_key())

        # ensure these fields don't carry stale _lock_btn attributes
        for w in (self.share, self.tb, self.cf, self.cp, self.rf, self.sum_field):
            if hasattr(w, "_lock_btn"):
                delattr(w, "_lock_btn")

        self.update_share_hint()
        self._resort()

    def _localized_title(self, lang: str) -> str:
        lang = (lang or "ru").lower()
        if lang.startswith("en"):
            part = "hot flow" if self.is_hot else "cold flow"
            return f"Component mixture ({part})"
        return QCoreApplication.translate(
            "MixPanel", "Смесь компонентов {part}"
        ).format(part=self._title_ru.lower())

    def retranslate_panel(self, lang: str) -> None:
        """Обновить надписи панели смеси при смене языка на лету."""
        try:
            self.box.setTitle(self._localized_title(lang))
            if lang.lower().startswith("en"):
                self.add_btn.setText("Add")
                self.lbl_share.setText("Share")
                self.lbl_sum.setText("Sum")
                self.rb_db.setText("Use parameters from NIST Chemistry WebBook")
                self.rb_manual.setText("Enter parameters manually")
                self.lbl_tb.setText("Boiling temperature, Tb [K]")
                self.lbl_cf.setText("Specific heat (liquid), C_f [kJ/kg·K]")
                self.lbl_cp.setText("Specific heat (vapor), C_p [kJ/kg·K]")
                self.lbl_rf.setText("Latent heat of phase change, r_f [kJ/kg]")
            else:
                self.add_btn.setText(self.box.tr("Добавить"))
                self.lbl_share.setText(self.box.tr("Доля"))
                self.lbl_sum.setText(self.box.tr("Сумма"))
                self.rb_db.setText(
                    self.box.tr("Взять параметры из справочника NIST Chemistry WebBook")
                )
                self.rb_manual.setText(self.box.tr("Ввести параметры вручную"))
                self.lbl_tb.setText(self.box.tr("Температура кипения, Tb  [ K ]"))
                self.lbl_cf.setText(
                    self.box.tr("Удельная теплоёмкость жидкости, C_f  [ кДж/кг·K ]")
                )
                self.lbl_cp.setText(
                    self.box.tr("Удельная теплоёмкость пара, C_p  [ кДж/кг·K ]")
                )
                self.lbl_rf.setText(
                    self.box.tr("Скрытая теплота фазового перехода, r_f  [ кДж/кг ]")
                )
        except Exception:
            pass

    def retranslate_existing_rows(self, lang: str) -> None:
        """Переименовать отображаемые имена компонентов в существующих строках под активный язык.
        Логические ключи компонентов остаются русскими; здесь меняется только DisplayRole.
        """
        try:
            to_en = (lang or "ru").lower().startswith("en")
        except Exception:
            to_en = False
        try:
            for r in range(self.model.rowCount()):
                idx = self.model.index(r, MixModel.COL_NAME)
                cur = str(self.model.data(idx, Qt.DisplayRole) or "")
                # Определим русский ключ для этого имени
                if cur in COMPONENT_DB:
                    ru_key = cur
                elif cur in COMPONENT_NAME_RU_FROM_EN:
                    ru_key = COMPONENT_NAME_RU_FROM_EN[cur]
                else:
                    # неизвестное имя — оставляем как есть
                    ru_key = cur
                new_disp = COMPONENT_NAME_EN.get(ru_key, ru_key) if to_en else ru_key
                if new_disp != cur:
                    self.model.setData(idx, new_disp, Qt.DisplayRole)
        except Exception:
            pass

    def widget(self) -> QGroupBox:
        return self.box

    def refresh_component_list(self) -> None:
        current_key = self._current_component_key()
        self._populate_component_combo(preserve_key=current_key)
        # обновить поля параметров (если режим DB)
        if self.rb_db.isChecked():
            self.fill_from_db(self._current_component_key())

    # сортировка по Tb автоматически
    def _resort(self) -> None:
        col = MixModel.COL_TB
        order = Qt.DescendingOrder if self.is_hot else Qt.AscendingOrder
        self.proxy.sort(col, order)

    def _on_model_changed(self, *args: Any) -> None:
        self.update_share_hint()
        self._resort()
        self._auto_export_csv()

    def _auto_export_csv(self) -> None:
        try:
            with open(self.export_path, "w", newline="", encoding="utf-8-sig") as f:
                wr = csv.writer(f, delimiter=";")
                wr.writerow(
                    [
                        self.box.tr("Компонент"),
                        self.box.tr("Доля"),
                        self.box.tr("Tb, K"),
                        self.box.tr("C_f, кДж/кг·K"),
                        self.box.tr("C_p, кДж/кг·K"),
                        self.box.tr("r_f, кДж/кг"),
                    ]
                )
                for r in range(self.model.rowCount()):
                    row = []
                    for c in range(self.model.columnCount()):
                        txt = (
                            self.model.data(self.model.index(r, c), Qt.DisplayRole)
                            or ""
                        )
                        if c != 0:
                            txt = txt.replace(".", ",")
                        row.append(txt)
                    wr.writerow(cast(List[str], row))
        except Exception:
            pass

    def current_sum(self) -> float:
        s = 0.0
        for r in range(self.model.rowCount()):
            txt = (
                self.model.data(self.model.index(r, MixModel.COL_SHARE), Qt.DisplayRole)
                or "0"
            )
            s += float(txt.replace(",", "."))
        return s

    def update_share_hint(self) -> None:
        total = self.current_sum()
        remaining = max(0.0, 1.0 - total)
        self.share.setPlaceholderText(f"≤ {remaining:.5f}")
        self.sum_field.setText(f"{total:.5f}")
        try:
            if abs(total - 1.0) <= 1e-4:
                # зелёный при корректной сумме
                self.sum_field.setStyleSheet("QLineEdit { background:#d9f7d9; }")
            else:
                # красный пока не 1.0
                self.sum_field.setStyleSheet("QLineEdit { background:#ffd6d6; }")
        except Exception:
            pass

    def on_mode_change(self, _checked: bool) -> None:
        manual = self.rb_manual.isChecked()
        for w in (self.tb, self.cf, self.cp, self.rf):
            set_enabled(w, manual)
        if self.rb_db.isChecked():
            self.fill_from_db(self._current_component_key())

        try:
            if manual:
                highlight_style = "QLineEdit { background: #fff7d6; }"
                for w in (self.tb, self.cf, self.cp, self.rf):
                    w.setStyleSheet(highlight_style)
            else:
                for w in (self.tb, self.cf, self.cp, self.rf):
                    if not w.isEnabled():
                        w.setStyleSheet("background:#f3f3f3;")
                    else:
                        w.setStyleSheet("")
        except Exception:
            pass

    def fill_from_db(self, name: str) -> None:
        # name ожидается как русский ключ
        props = COMPONENT_DB.get(name)
        if props:
            tb, cf, cp, rf = props
            self.tb.setText(f"{tb}")
            self.cf.setText(f"{cf}")
            self.cp.setText(f"{cp}")
            self.rf.setText(f"{rf}")
        else:
            for w in (self.tb, self.cf, self.cp, self.rf):
                w.setText("0.0")

    def _is_lang_en(self) -> bool:
        try:
            app_inst = QApplication.instance()
            active_lang = (
                str(getattr(app_inst, "_app_translator_lang", "") or "").lower()
                if app_inst is not None
                else ""
            )
            desired = str(QSettings().value("ui/language", "ru") or "ru").lower()
            return desired == "en" or active_lang.startswith("en")
        except Exception:
            return False

    def _current_component_key(self) -> str:
        try:
            ru = self.comp.currentData()
            if isinstance(ru, str) and ru in COMPONENT_DB:
                return ru
        except Exception:
            pass
        try:
            disp = self.comp.currentText()
            if disp in COMPONENT_DB:
                return disp
            if disp in COMPONENT_NAME_RU_FROM_EN:
                return COMPONENT_NAME_RU_FROM_EN[disp]
        except Exception:
            pass
        # fallback: первый ключ
        try:
            return next(iter(sorted(COMPONENT_DB.keys())))
        except Exception:
            return ""

    def _populate_component_combo(self, preserve_key: Optional[str] = None) -> None:
        try:
            en = self._is_lang_en()
            self.comp.blockSignals(True)
            self.comp.clear()
            for ru_name in sorted(COMPONENT_DB.keys()):
                disp = COMPONENT_NAME_EN.get(ru_name, ru_name) if en else ru_name
                self.comp.addItem(disp, ru_name)
            # восстановить выбор
            key = preserve_key or (
                self.comp.itemData(0) if self.comp.count() > 0 else None
            )
            if isinstance(key, str):
                idx = self.comp.findData(key)
                if idx >= 0:
                    self.comp.setCurrentIndex(idx)
                elif self.comp.count() > 0:
                    self.comp.setCurrentIndex(0)
        except Exception:
            pass
        finally:
            try:
                self.comp.blockSignals(False)
            except Exception:
                pass

    def validate_share_max1(self) -> None:
        val = to_float(self.share.text())
        if val > 1.0:
            QMessageBox.warning(
                self.box,
                self.box.tr("Доля"),
                self.box.tr("Доля компонента не может превышать 1. Повторите ввод."),
            )
            self.share.clear()
            self.share.setFocus()

    def on_add(self) -> None:
        remaining = max(0.0, 1.0 - self.current_sum())
        share_val = to_float(self.share.text())
        if share_val > 1.0 + 1e-12:
            QMessageBox.warning(
                self.box,
                "Доля",
                "Доля компонента не может превышать 1. Повторите ввод.",
            )
            self.share.clear()
            self.share.setFocus()
            return
        if share_val <= 0.0:
            QMessageBox.warning(
                self.box,
                self.box.tr("Доля"),
                self.box.tr("Введите положительную долю > 0."),
            )
            return
        if share_val > remaining + 1e-12:
            if remaining <= 0.0:
                QMessageBox.warning(
                    self.box,
                    self.box.tr("Сумма долей"),
                    self.box.tr("Сумма долей уже равна 1.0."),
                )
                return
            share_val = remaining
            self.share.setText(f"{share_val:.5f}")
        # Всегда используем русский ключ для доступа к БД, а отображаемое имя — по активному языку
        ru_key = self._current_component_key()
        en_mode = self._is_lang_en()
        display_name = COMPONENT_NAME_EN.get(ru_key, ru_key) if en_mode else ru_key
        if self.rb_db.isChecked():
            tb, cf, cp, rf = COMPONENT_DB.get(ru_key, (0.0, 0.0, 0.0, 0.0))
        else:
            tb, cf, cp, rf = (
                to_float(self.tb.text()),
                to_float(self.cf.text()),
                to_float(self.cp.text()),
                to_float(self.rf.text()),
            )
        self.model.add_or_update(display_name, share_val, tb, cf, cp, rf)
        self.share.clear()

    def ask_delete(self, count: int) -> bool:
        box = QMessageBox(self.box)
        box.setIcon(QMessageBox.Question)
        # Заголовок и текст: при EN задаём явные английские строки, иначе через tr()
        if self._is_lang_en():
            title = "Delete"
            text = f"Delete {count} row(s)?"
        else:
            title = self.box.tr("Удаление")
            text = self.box.tr("Удалить {n} строку(и)?").format(n=count)
        box.setWindowTitle(title)
        box.setText(text)
        # Кнопки: для режима EN ставим явные английские подписи, иначе — через tr()
        if self._is_lang_en():
            yes_btn = box.addButton("Yes", QMessageBox.AcceptRole)
            no_btn = box.addButton("No", QMessageBox.RejectRole)
        else:
            yes_btn = box.addButton(self.box.tr("Да"), QMessageBox.AcceptRole)
            no_btn = box.addButton(self.box.tr("Нет"), QMessageBox.RejectRole)
        box.setDefaultButton(no_btn)  # по умолчанию Нет
        box.exec_()
        return box.clickedButton() is yes_btn

    def selected_source_rows(self) -> List[int]:
        rows: List[int] = []
        for proxy_index in self.view.selectionModel().selectedRows():
            src_idx = self.proxy.mapToSource(proxy_index)
            rows.append(src_idx.row())
        return sorted(set(rows), reverse=True)

    def delete_selected_rows(self) -> None:
        rows = self.selected_source_rows()
        if not rows:
            if self._is_lang_en():
                QMessageBox.information(self.box, "Delete", "Select row(s) to delete.")
            else:
                QMessageBox.information(
                    self.box,
                    self.box.tr("Удаление"),
                    self.box.tr("Выберите строку(и) для удаления."),
                )
            return
        if not self.ask_delete(len(rows)):
            return
        self.model.remove_rows(rows)

    def on_double_click(self, index: QModelIndex) -> None:
        if not index.isValid():
            return
        if not self.ask_delete(1):
            return
        self.model.removeRow(self.proxy.mapToSource(index).row())

    def mix_rows(self) -> List[MixRow]:
        return self.model.rows_as_dicts()


# ===================== ПАНЕЛЬ ГИДРОДИНАМИКИ =====================
class HydroPanel(QGroupBox):
    def __init__(
        self, title: str = "Гидродинамика потоков", parent: Optional[QWidget] = None
    ):
        super().__init__(title, parent)
        # Расширяемая по ширине, высота по содержимому
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        base = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "assets", "images"
        )
        self._images = {
            "mix_mix": os.path.join(base, "one.png"),
            "parallel": os.path.join(base, "two.png"),
            "mix_cold_disp_hot": os.path.join(base, "three.png"),
            "mix_hot_disp_cold": os.path.join(base, "four.png"),
            "counter": os.path.join(base, "five.png"),
        }
        root = QHBoxLayout(self)
        root.setContentsMargins(6, 6, 6, 6)
        root.setSpacing(6)
        left = QVBoxLayout()
        right = QVBoxLayout()
        left.setContentsMargins(4, 4, 4, 4)
        right.setContentsMargins(4, 4, 4, 4)
        root.addLayout(left)
        root.addLayout(right)
        root.setStretch(0, 0)
        root.setStretch(1, 0)

        self.rb_mix_mix = QRadioButton(self.tr("Смешение - смешение"))
        self.rb_parallel = QRadioButton(self.tr("Вытеснение - вытеснение (прямоток)"))
        self.rb_mix_cold = QRadioButton(self.tr("Смешение (хол.) - вытеснение (гор.)"))
        self.rb_mix_hot = QRadioButton(self.tr("Смешение (гор.) - вытеснение (хол.)"))
        self.rb_counter = QRadioButton(self.tr("Вытеснение - вытеснение (противоток)"))
        for rb in (
            self.rb_mix_mix,
            self.rb_parallel,
            self.rb_mix_cold,
            self.rb_mix_hot,
            self.rb_counter,
        ):
            left.addWidget(rb)
        left.addStretch(1)

        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        # Фиксированный размер изображения
        self.image_label.setFixedSize(350, 175)
        self.image_label.setFrameShape(QFrame.Box)
        self.image_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        right.addWidget(self.image_label, 0, Qt.AlignTop | Qt.AlignHCenter)

        def _on_rb_mix_mix_toggled(on: bool) -> None:
            if on:
                self._set_mode("mix_mix")

        def _on_rb_parallel_toggled(on: bool) -> None:
            if on:
                self._set_mode("parallel")

        def _on_rb_mix_cold_toggled(on: bool) -> None:
            if on:
                self._set_mode("mix_cold_disp_hot")

        def _on_rb_mix_hot_toggled(on: bool) -> None:
            if on:
                self._set_mode("mix_hot_disp_cold")

        def _on_rb_counter_toggled(on: bool) -> None:
            if on:
                self._set_mode("counter")

        self.rb_mix_mix.toggled.connect(_on_rb_mix_mix_toggled)
        self.rb_parallel.toggled.connect(_on_rb_parallel_toggled)
        self.rb_mix_cold.toggled.connect(_on_rb_mix_cold_toggled)
        self.rb_mix_hot.toggled.connect(_on_rb_mix_hot_toggled)
        self.rb_counter.toggled.connect(_on_rb_counter_toggled)

        self.rb_mix_mix.setChecked(True)
        self._set_mode("mix_mix")

    def retranslate_panel(self, lang: str) -> None:
        try:
            if lang.startswith("en"):
                self.setTitle("Flow hydrodynamics")
                self.rb_mix_mix.setText("Mixing - mixing")
                self.rb_parallel.setText("Displacement - displacement (cocurrent)")
                self.rb_mix_cold.setText("Mixing (cold) - displacement (hot)")
                self.rb_mix_hot.setText("Mixing (hot) - displacement (cold)")
                self.rb_counter.setText("Displacement - displacement (countercurrent)")
            else:
                self.setTitle(self.tr("Гидродинамика потоков"))
                self.rb_mix_mix.setText(self.tr("Смешение - смешение"))
                self.rb_parallel.setText(self.tr("Вытеснение - вытеснение (прямоток)"))
                self.rb_mix_cold.setText(self.tr("Смешение (хол.) - вытеснение (гор.)"))
                self.rb_mix_hot.setText(self.tr("Смешение (гор.) - вытеснение (хол.)"))
                self.rb_counter.setText(self.tr("Вытеснение - вытеснение (противоток)"))
        except Exception:
            pass

    def current_schema(self) -> str:
        """Возвращает идентификатор схемы (Schema1..Schema5) согласно выбранной радиокнопке."""
        if self.rb_mix_mix.isChecked():
            return "Schema1"
        if self.rb_parallel.isChecked():
            return "Schema2"
        if self.rb_mix_cold.isChecked():
            return "Schema3"
        if self.rb_mix_hot.isChecked():
            return "Schema4"
        if self.rb_counter.isChecked():
            return "Schema5"
        return "Schema1"

    def _set_mode(self, key: str) -> None:
        pix = QPixmap(self._images.get(key, ""))
        if pix.isNull():
            self.image_label.setText(self.tr("Нет изображения"))
            self.image_label.setPixmap(QPixmap())
            return
        self.image_label.setPixmap(
            pix.scaled(
                self.image_label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation
            )
        )

    def resizeEvent(self, e: QEvent) -> None:  # type: ignore[override]
        super().resizeEvent(e)  # type: ignore[arg-type]
        if self.rb_mix_mix.isChecked():
            self._set_mode("mix_mix")
        elif self.rb_parallel.isChecked():
            self._set_mode("parallel")
        elif self.rb_mix_cold.isChecked():
            self._set_mode("mix_cold_disp_hot")
        elif self.rb_mix_hot.isChecked():
            self._set_mode("mix_hot_disp_cold")
        elif self.rb_counter.isChecked():
            self._set_mode("counter")


# ===================== ПАНЕЛЬ ВЫХОДНЫХ ПАРАМЕТРОВ =====================
class OutputPanel(QGroupBox):
    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(
            QCoreApplication.translate("OutputPanel", "Параметры теплообменника"),
            parent,
        )
        # Расширяемая по ширине, фиксированная по высоте
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        g = QGridLayout(self)
        self.q = num_edit(read_only=False, fixed_width=120)
        self.q_lock = lock_button_for(self.q)
        self.sigma = num_edit(read_only=True, fixed_width=120)
        self.k = num_edit(read_only=True, fixed_width=120)
        # стартовые значения
        self.sigma.setText("0.0")
        set_enabled(self.sigma, False)
        self.k.setText("0.0")
        set_enabled(self.k, False)
        self.lbl_q = QLabel(self.tr("Тепловая нагрузка, Q [кВт]"))
        g.addWidget(self.lbl_q, 0, 0)
        hq = QHBoxLayout()
        hq.setContentsMargins(0, 0, 0, 0)
        hq.addWidget(self.q)
        hq.addWidget(self.q_lock)
        g.addLayout(hq, 0, 1)
        self.lbl_sigma = QLabel(self.tr("Производство энтропии, σ [кВт/К]"))
        g.addWidget(self.lbl_sigma, 1, 0)
        g.addWidget(self.sigma, 1, 1)
        self.lbl_k = QLabel(self.tr("Коэффициент теплопередачи, K [кВт/К]"))
        g.addWidget(self.lbl_k, 2, 0)
        g.addWidget(self.k, 2, 1)
        # Removed schema_label (schema info now only in status bar)
        # remove stale lock attributes if any
        for w in (self.sigma, self.k):
            if hasattr(w, "_lock_btn"):
                delattr(w, "_lock_btn")
        # auto-disable Q after editingFinished
        try:
            self.q.editingFinished.connect(auto_disable_handler(self.q))
        except Exception:
            pass

    def clear_values(self) -> None:
        for w in (self.q,):
            w.clear()
        # сохранить sigma/k как read-only 0.0
        self.sigma.setText("0.0")
        set_enabled(self.sigma, False)
        self.k.setText("0.0")
        set_enabled(self.k, False)

    # schema label removed

    def retranslate_panel(self, lang: str) -> None:
        try:
            if lang.startswith("en"):
                self.setTitle("Heat exchanger parameters")
                self.lbl_q.setText("Heat load, Q [kW]")
                self.lbl_sigma.setText("Entropy production, σ [kW/K]")
                self.lbl_k.setText("Heat transfer coefficient, K [kW/K]")
            else:
                self.setTitle(self.tr("Параметры теплообменника"))
                self.lbl_q.setText(self.tr("Тепловая нагрузка, Q [кВт]"))
                self.lbl_sigma.setText(self.tr("Производство энтропии, σ [кВт/К]"))
                self.lbl_k.setText(self.tr("Коэффициент теплопередачи, K [кВт/К]"))
        except Exception:
            pass


# ===================== ГЛАВНОЕ ОКНО =====================
class MainWindow(QMainWindow):
    def __init__(self, initial_theme: str = "system", initial_language: str = "ru"):
        super().__init__()
        # flag to indicate we are currently importing data (suppress full on_calc triggers)
        self._importing = False
        # after importing, suppress full sigma/K calculation on schema toggle until user presses Calculate
        self._suppress_full_calc_after_import = False
        # track changes after import to show recalc button
        self._post_import_changed = False
        self.setWindowTitle(self.tr("Двухпоточный теплообмен"))
        # Сделаем окно ресайзабельным: установим минимальный размер
        # и стартовый размер. Высоту уменьшаем на 50px (с 1025 до 975).
        self.setMinimumSize(1600, 975)
        self.resize(1600, 975)
        # Центрирование окна при первом запуске
        try:
            scr = self.screen()  # type: ignore[attr-defined]
        except Exception:
            scr = None
        if scr is None:
            try:
                scr = QApplication.primaryScreen()
            except Exception:
                scr = None
        if scr is not None:
            try:
                geo = scr.availableGeometry()
                x = geo.x() + (geo.width() - self.width()) // 2
                y = geo.y() + (geo.height() - self.height()) // 2
                self.move(max(geo.left(), x), max(geo.top(), y))
            except Exception:
                pass
        # статусная строка
        self.status = self.statusBar()
        try:
            self.status.showMessage(self.tr("Готово"))
        except Exception:
            pass
        # флаг: было ли явное нажатие кнопки Вычислить после последнего изменения схемы/сброса
        self._explicit_calc_done = False
        self._results_stale = False

        # File menu: Import/Export inputs (JSON)
        try:
            # Язык интерфейса зафиксирован на русском

            self.file_menu = self.menuBar().addMenu(self.tr("Файл"))
            self.act_imp_inputs = QAction(self.tr("Импорт входных данных..."), self)
            self.act_exp_inputs = QAction(self.tr("Экспорт входных данных..."), self)
            self.act_imp_inputs_xlsx = QAction(
                self.tr("Импорт из Excel (.xlsx)..."), self
            )
            self.act_exp_inputs_xlsx = QAction(
                self.tr("Экспорт в Excel (.xlsx)..."), self
            )
            # Действия с базой компонентов
            self.act_imp_db = QAction(
                self.tr("Импорт базы компонентов (Excel)..."), self
            )
            # Убрали CSV-экспорт базы компонентов, оставляем только Excel-экспорт
            self.act_exp_db_xlsx = QAction(
                self.tr("Экспорт базы компонентов (Excel)..."), self
            )
            self.file_menu.addAction(self.act_imp_inputs)
            self.file_menu.addAction(self.act_exp_inputs)
            self.file_menu.addAction(self.act_imp_inputs_xlsx)
            self.file_menu.addAction(self.act_exp_inputs_xlsx)
            self.file_menu.addSeparator()
            self.file_menu.addAction(self.act_imp_db)
            self.file_menu.addAction(self.act_exp_db_xlsx)

            # Меню "Вид"
            self.view_menu = self.menuBar().addMenu(self.tr("Вид"))
            self.act_reset_view = QAction(self.tr("Сбросить вид (по умолчанию)"), self)
            self.view_menu.addAction(self.act_reset_view)
            # --- Тема ---
            self.theme_menu = self.view_menu.addMenu(self.tr("Тема"))
            self._theme_group = QActionGroup(self)
            self._theme_group.setExclusive(True)
            self._act_theme_light = QAction(self.tr("Светлая"), self)
            self._act_theme_light.setCheckable(True)
            self._act_theme_dark = QAction(self.tr("Тёмная"), self)
            self._act_theme_dark.setCheckable(True)
            self._act_theme_system = QAction(self.tr("Системная"), self)
            self._act_theme_system.setCheckable(True)
            for a in (
                self._act_theme_light,
                self._act_theme_dark,
                self._act_theme_system,
            ):
                self._theme_group.addAction(a)
                self.theme_menu.addAction(a)
            # --- Язык (устаревшее меню, отключено по запросу пользователя; оставить на будущее) ---
            if SHOW_LANGUAGE_MENU:
                self.lang_menu = self.view_menu.addMenu(self.tr("Язык"))
                self._lang_group = QActionGroup(self)
                self._lang_group.setExclusive(True)
                self._act_lang_ru = QAction(self.tr("Русский"), self)
                self._act_lang_ru.setCheckable(True)
                self._act_lang_en = QAction(self.tr("English"), self)
                self._act_lang_en.setCheckable(True)
                for a in (self._act_lang_ru, self._act_lang_en):
                    self._lang_group.addAction(a)
                    self.lang_menu.addAction(a)
                self._act_lang_ru.triggered.connect(
                    lambda: self._on_language_selected("ru")
                )
                self._act_lang_en.triggered.connect(
                    lambda: self._on_language_selected("en")
                )
            self.act_imp_inputs.triggered.connect(self.import_inputs)  # type: ignore[call-arg]
            self.act_exp_inputs.triggered.connect(self.export_inputs)  # type: ignore[call-arg]
            self.act_imp_inputs_xlsx.triggered.connect(self.import_inputs_xlsx)  # type: ignore[call-arg]
            self.act_exp_inputs_xlsx.triggered.connect(self.export_inputs_xlsx)  # type: ignore[call-arg]
            self.act_imp_db.triggered.connect(self.import_component_db_xlsx)  # type: ignore[call-arg]
            self.act_exp_db_xlsx.triggered.connect(self.export_component_db_xlsx)  # type: ignore[call-arg]
            # --- Меню помощи ---
            self.help_menu = self.menuBar().addMenu(self.tr("Помощь"))
            self.act_help = QAction(self.tr("Справка"), self)
            self.act_logs = QAction(self.tr("Логи"), self)
            self.act_about = QAction(self.tr("О программе"), self)
            self.act_license = QAction(self.tr("Лицензионное соглашение"), self)
            self.help_menu.addAction(self.act_help)
            self.help_menu.addAction(self.act_logs)
            self.help_menu.addSeparator()
            self.help_menu.addAction(self.act_license)
            self.help_menu.addAction(self.act_about)
            self.act_help.triggered.connect(self.show_help_dialog)
            self.act_logs.triggered.connect(self.show_logs_dialog)
            self.act_license.triggered.connect(self.show_license_dialog)
            self.act_about.triggered.connect(self.show_about_dialog)
            self.act_reset_view.triggered.connect(self.reset_view)
            # обработчики темы (языка нет)
            self._act_theme_light.triggered.connect(
                lambda: self._on_theme_selected("light")
            )
            self._act_theme_dark.triggered.connect(
                lambda: self._on_theme_selected("dark")
            )
            self._act_theme_system.triggered.connect(
                lambda: self._on_theme_selected("system")
            )
            # Применим заголовки главного меню согласно выбранному языку
            try:
                self._apply_menu_language(initial_language)
            except Exception:
                pass
        except Exception as e:
            logger.exception("Ошибка создания меню: %s", e)

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout()
        central.setLayout(layout)

        # Верхняя панель с кнопками RU/EN (справа)
        topbar = QHBoxLayout()
        topbar.setContentsMargins(0, 0, 0, 0)
        topbar.setSpacing(6)
        topbar.addStretch(1)
        self.lang_btn_ru = QPushButton("RU")
        self.lang_btn_en = QPushButton("EN")
        for b in (self.lang_btn_ru, self.lang_btn_en):
            try:
                b.setFixedHeight(24)
                b.setCheckable(True)
                b.setCursor(Qt.PointingHandCursor)
            except Exception:
                pass
        # подсветка выбранной кнопки
        try:
            sel_style = (
                "QPushButton { padding: 4px 10px; } "
                "QPushButton:checked { background: #2d7dff; color: white; font-weight: 600; }"
            )
            self.lang_btn_ru.setStyleSheet(sel_style)
            self.lang_btn_en.setStyleSheet(sel_style)
        except Exception:
            pass
        # эксклюзивная группа
        try:
            self._lang_btn_group = QButtonGroup(self)
            self._lang_btn_group.setExclusive(True)
            self._lang_btn_group.addButton(self.lang_btn_ru)
            self._lang_btn_group.addButton(self.lang_btn_en)
        except Exception:
            pass
        try:
            self.lang_btn_ru.setToolTip(self.tr("Язык: русский"))
            self.lang_btn_en.setToolTip(self.tr("Язык: английский"))
        except Exception:
            pass
        try:
            self.lang_btn_ru.clicked.connect(lambda: self._on_language_selected("ru"))
            self.lang_btn_en.clicked.connect(lambda: self._on_language_selected("en"))
        except Exception:
            pass
        topbar.addWidget(self.lang_btn_ru)
        topbar.addWidget(self.lang_btn_en)
        layout.addLayout(topbar)

        # потоки
        row1 = QHBoxLayout()
        layout.addLayout(row1)
        self.cold_panel = FlowPanel("Холодный поток", sign="−")
        self.hot_panel = FlowPanel("Горячий поток", sign="+")
        row1.addWidget(self.cold_panel.widget())
        row1.addWidget(self.hot_panel.widget())

        # смеси
        row2 = QHBoxLayout()
        layout.addLayout(row2)
        self.cold_mix = MixPanel(
            "холодного потока",
            is_hot=False,
            export_path=os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "data",
                "csv",
                "cold_mix.csv",
            ),
        )
        self.hot_mix = MixPanel(
            "горячего потока",
            is_hot=True,
            export_path=os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "data", "csv", "hot_mix.csv"
            ),
        )
        row2.addWidget(self.cold_mix.widget())
        row2.addWidget(self.hot_mix.widget())

        # connect mix model changes to update button state and attempt minimal auto-calc
        try:
            self.cold_mix.model.dataChanged.connect(self._on_mix_changed)  # type: ignore[call-arg]
            self.cold_mix.model.rowsInserted.connect(self._on_mix_changed)  # type: ignore[call-arg]
            self.cold_mix.model.rowsRemoved.connect(self._on_mix_changed)  # type: ignore[call-arg]
            self.hot_mix.model.dataChanged.connect(self._on_mix_changed)  # type: ignore[call-arg]
            self.hot_mix.model.rowsInserted.connect(self._on_mix_changed)  # type: ignore[call-arg]
            self.hot_mix.model.rowsRemoved.connect(self._on_mix_changed)  # type: ignore[call-arg]
        except Exception:
            pass

        # гидродинамика + правый столбец (OutputPanel + кнопки)
        row3 = QHBoxLayout()
        row3.setSpacing(12)
        layout.addLayout(row3)
        self.hydro = HydroPanel()
        row3.addWidget(self.hydro, 1)

        # правая колонка: OutputPanel + кнопки (чуть шире)
        right_col = QVBoxLayout()
        right_col.setContentsMargins(0, 0, 0, 0)
        right_col.setSpacing(8)
        self.out_panel = OutputPanel()
        self.out_panel.setMinimumWidth(750)
        right_col.addWidget(self.out_panel)
        btns = QVBoxLayout()
        btns.setContentsMargins(0, 0, 0, 0)
        btns.setSpacing(8)
        # Кнопки справа (язык интерфейса зафиксирован на русском)
        self.calc_btn = QPushButton(self.tr("Вычислить"))
        self.reset_btn = QPushButton(self.tr("Очистить параметры"))
        self.calc_btn.setMinimumHeight(36)
        self.reset_btn.setMinimumHeight(36)
        self.analysis_btn = QPushButton(self.tr("Провести анализ"))
        self.analysis_btn.setToolTip("Провести анализ изменяя доли компонентов потоков")
        self.analysis_btn.setMinimumHeight(36)
        self.recalc_btn = QPushButton(self.tr("Перерасчёт"))
        self.recalc_btn.setToolTip("Пересчитать после изменений")
        self.recalc_btn.setMinimumHeight(36)
        self.recalc_btn.hide()
        btns.addWidget(self.calc_btn)
        btns.addWidget(self.recalc_btn)
        btns.addWidget(self.analysis_btn)
        btns.addWidget(self.reset_btn)
        right_col.addLayout(btns)
        # таймер мигания для кнопки перерасчёта
        from PyQt5.QtCore import QTimer as _QTimer

        self._recalc_blink_state = False

        def _blink_recalc():
            try:
                if not self.recalc_btn.isVisible():
                    # вернуть стандартный стиль
                    self.recalc_btn.setStyleSheet("")
                    return
                self._recalc_blink_state = not getattr(
                    self, "_recalc_blink_state", False
                )
                if self._recalc_blink_state:
                    self.recalc_btn.setStyleSheet(
                        "QPushButton { background:#ff4d4f; color:white; font-weight:bold; }"
                    )
                else:
                    self.recalc_btn.setStyleSheet(
                        "QPushButton { background:#ffcccc; color:#800; font-weight:bold; }"
                    )
            except Exception:
                pass

        self._recalc_blink_timer = _QTimer(self)
        self._recalc_blink_timer.timeout.connect(_blink_recalc)  # type: ignore
        self._recalc_blink_timer.start(700)
        right_col.addStretch(1)
        row3.addLayout(right_col, 0)
        # Обе колонки расширяются равномерно по ширине, как панели смесей
        row3.setStretch(0, 1)  # HydroPanel
        row3.setStretch(1, 1)  # Right column

        # связи
        try:
            # Explicit user click should clear import suppression and run full calc
            self.calc_btn.clicked.connect(self._on_calc_button_clicked)
        except Exception:
            pass
        try:
            self.recalc_btn.clicked.connect(self._on_recalc_clicked)
        except Exception:
            pass
        try:
            self.reset_btn.clicked.connect(self.on_reset)
        except Exception:
            pass
        try:
            self.analysis_btn.clicked.connect(self.open_analysis_window)
        except Exception:
            pass
        # автопересчёт при смене схемы если есть пред. результат
        try:
            for rb in (
                self.hydro.rb_mix_mix,
                self.hydro.rb_parallel,
                self.hydro.rb_mix_cold,
                self.hydro.rb_mix_hot,
                self.hydro.rb_counter,
            ):
                rb.toggled.connect(self._on_schema_changed)  # type: ignore[call-arg]
        except Exception:
            pass
        # взаимная блокировка: ввод Q блокирует T+out и наоборот
        try:
            # используем editingFinished — расчёт выполняется после завершения ввода
            self.out_panel.q.editingFinished.connect(self._on_q_edit_finished)  # type: ignore[call-arg]
            self.hot_panel.t_out.editingFinished.connect(
                self._on_tplus_out_edit_finished  # type: ignore[call-arg]
            )
            # авто-блокировка этих полей после редактирования
            self.out_panel.q.editingFinished.connect(
                auto_disable_handler(self.out_panel.q)
            )
            self.hot_panel.t_out.editingFinished.connect(
                auto_disable_handler(self.hot_panel.t_out)
            )
        except Exception:
            pass
        # Подключим auto-calc при завершении ввода основных входных полей
        try:
            for w in (
                self.cold_panel.t_in,
                self.cold_panel.t_out,
                self.cold_panel.m,
                self.hot_panel.t_in,
                self.hot_panel.t_out,
                self.hot_panel.m,
            ):
                w.editingFinished.connect(self._normalize_input)
        except Exception:
            pass
        # Also trigger auto-calc when mixtures change (components/dolya edited or rows changed)
        # Automatic recalculation on every mix change was removed to avoid noisy updates.

        # update calc button appearance when inputs change
        try:
            widgets_and_models = [
                self.cold_panel.t_in,
                self.cold_panel.t_out,
                self.cold_panel.m,
                self.hot_panel.t_in,
                self.hot_panel.t_out,
                self.hot_panel.m,
                self.out_panel.q,
            ]
            for w in widgets_and_models:
                # connect editingFinished if available
                if hasattr(w, "editingFinished"):
                    try:
                        w.editingFinished.connect(self._update_calc_button_state)  # type: ignore[call-arg]
                        # При любом завершении ввода — инициируем режим «перерасчёта»: показать кнопку и обновить значения
                        w.editingFinished.connect(self._on_any_input_changed)  # type: ignore[call-arg]
                        # Добавим фильтр на уход фокуса для надёжной авто-блокировки и показа «Перерасчёта»
                        try:
                            _f = AutoLockRecalcFilter(
                                self, w, self._on_any_input_changed
                            )
                            setattr(w, "_autoLockFilter", _f)
                            w.installEventFilter(_f)
                        except Exception:
                            pass
                    except Exception:
                        pass
                # connect dataChanged if available (models)
                if hasattr(w, "dataChanged"):
                    try:
                        w.dataChanged.connect(self._update_calc_button_state)  # type: ignore[call-arg]
                    except Exception:
                        pass
        except Exception:
            pass

        # initial update of button state
        self._update_calc_button_state()

        # Подключим слежение за изменениями полей потоков для пометки устаревания результатов
        try:

            def _mark_change():
                try:
                    self._mark_stale_results()
                except Exception:
                    pass

            for w in (
                self.cold_panel.t_in,
                self.cold_panel.t_out,
                self.cold_panel.m,
                self.cold_panel.p,
                self.hot_panel.t_in,
                self.hot_panel.t_out,
                self.hot_panel.m,
                self.hot_panel.p,
                self.out_panel.q,
            ):
                try:
                    w.editingFinished.connect(_mark_change)  # type: ignore
                except Exception:
                    pass
        except Exception:
            pass

        # Снять стартовый фокус с поля T_in холодного потока (курсор не должен мигать там при запуске)
        try:
            QTimer.singleShot(0, self._remove_initial_focus)  # type: ignore[arg-type]
        except Exception:
            pass

        # --- Применяем начальные тема/язык ---
        try:
            self._init_theme_language(initial_theme, initial_language)
        except Exception:
            pass

        # Применим переводы к элементам UI при старте, чтобы покрыть случаи неполного .qm
        try:
            lang0 = (initial_language or "ru").lower()
            # меню
            try:
                self._apply_menu_language(lang0)
            except Exception:
                pass
            # списки компонентов + панели
            try:
                self.cold_mix.refresh_component_list()
                self.hot_mix.refresh_component_list()
            except Exception:
                pass
            try:
                if lang0.startswith("en"):
                    self.setWindowTitle("Two-stream heat exchanger")
                else:
                    self.setWindowTitle(self.tr("Двухпоточный теплообмен"))
            except Exception:
                pass
            try:
                self.cold_panel.retranslate_panel(lang0)
                self.hot_panel.retranslate_panel(lang0)
            except Exception:
                pass
            try:
                self.hydro.retranslate_panel(lang0)
            except Exception:
                pass
            try:
                self.out_panel.retranslate_panel(lang0)
            except Exception:
                pass
            try:
                self.cold_mix.retranslate_panel(lang0)
                self.hot_mix.retranslate_panel(lang0)
            except Exception:
                pass
            # Заголовки таблиц
            try:
                self.cold_mix.model.retranslate_headers()
                self.hot_mix.model.retranslate_headers()
            except Exception:
                pass
            # Переименовать уже добавленные строки смесей под активный язык
            try:
                self.cold_mix.retranslate_existing_rows(lang0)
                self.hot_mix.retranslate_existing_rows(lang0)
            except Exception:
                pass
            # Кнопки справа
            try:
                if lang0.startswith("en"):
                    self.calc_btn.setText("Calculate")
                    self.reset_btn.setText("Clear parameters")
                    self.analysis_btn.setText("Run analysis")
                    self.analysis_btn.setToolTip(
                        "Run analysis by varying component shares"
                    )
                    self.recalc_btn.setText("Recalculate")
                    self.recalc_btn.setToolTip("Recalculate after changes")
                else:
                    self.calc_btn.setText(self.tr("Вычислить"))
                    self.reset_btn.setText(self.tr("Очистить параметры"))
                    self.analysis_btn.setText(self.tr("Провести анализ"))
                    self.analysis_btn.setToolTip(
                        self.tr("Провести анализ изменяя доли компонентов потоков")
                    )
                    self.recalc_btn.setText(self.tr("Перерасчёт"))
                    self.recalc_btn.setToolTip(self.tr("Пересчитать после изменений"))
            except Exception:
                pass
            # Обновим подсветку RU/EN кнопок в топбаре
            try:
                if lang0.startswith("en"):
                    self.lang_btn_en.setChecked(True)
                    self.lang_btn_ru.setChecked(False)
                else:
                    self.lang_btn_ru.setChecked(True)
                    self.lang_btn_en.setChecked(False)
            except Exception:
                pass
            # Текст статус-бара
            try:
                if lang0.startswith("en"):
                    self.status.showMessage("Ready")
                else:
                    self.status.showMessage(self.tr("Готово"))
            except Exception:
                pass
        except Exception:
            pass

    # ============== Тема и язык ==============
    def _init_theme_language(self, theme: str, lang: str) -> None:
        # Тема
        theme = (theme or "system").lower()
        if theme not in ("light", "dark", "system"):
            theme = "system"
        self.apply_theme(theme)
        if theme == "light":
            self._act_theme_light.setChecked(True)
        elif theme == "dark":
            self._act_theme_dark.setChecked(True)
        else:
            self._act_theme_system.setChecked(True)
        # Язык — отмечаем текущий выбор (кнопки в правом верхнем углу +, при включённом флаге, элементы меню)
        try:
            lang_norm = (lang or "ru").lower()
            if lang_norm.startswith("en"):
                try:
                    if SHOW_LANGUAGE_MENU:
                        self._act_lang_en.setChecked(True)
                except Exception:
                    pass
                try:
                    self.lang_btn_en.setChecked(True)
                    self.lang_btn_ru.setChecked(False)
                except Exception:
                    pass
            else:
                try:
                    if SHOW_LANGUAGE_MENU:
                        self._act_lang_ru.setChecked(True)
                except Exception:
                    pass
                try:
                    self.lang_btn_ru.setChecked(True)
                    self.lang_btn_en.setChecked(False)
                except Exception:
                    pass
        except Exception:
            pass

    def _apply_menu_language(self, lang: str) -> None:
        """Применить язык к заголовкам главного меню без использования Qt-переводчика."""
        lang_norm = (lang or "ru").lower()
        try:
            self.file_menu.setTitle(
                "File" if lang_norm.startswith("en") else self.tr("Файл")
            )
        except Exception:
            pass
        # Пункты File
        try:
            if lang_norm.startswith("en"):
                self.act_imp_inputs.setText("Import inputs…")
                self.act_exp_inputs.setText("Export inputs…")
                self.act_imp_inputs_xlsx.setText("Import from Excel (.xlsx)…")
                self.act_exp_inputs_xlsx.setText("Export to Excel (.xlsx)…")
                self.act_imp_db.setText("Import component DB (Excel)…")
                self.act_exp_db_xlsx.setText("Export component DB (Excel)…")
            else:
                self.act_imp_inputs.setText(self.tr("Импорт входных данных..."))
                self.act_exp_inputs.setText(self.tr("Экспорт входных данных..."))
                self.act_imp_inputs_xlsx.setText(self.tr("Импорт из Excel (.xlsx)..."))
                self.act_exp_inputs_xlsx.setText(self.tr("Экспорт в Excel (.xlsx)..."))
                self.act_imp_db.setText(self.tr("Импорт базы компонентов (Excel)..."))
                self.act_exp_db_xlsx.setText(
                    self.tr("Экспорт базы компонентов (Excel)...")
                )
        except Exception:
            pass
        try:
            self.view_menu.setTitle(
                "View" if lang_norm.startswith("en") else self.tr("Вид")
            )
        except Exception:
            pass
        # Пункты View
        try:
            self.theme_menu.setTitle(
                "Theme" if lang_norm.startswith("en") else self.tr("Тема")
            )
            if SHOW_LANGUAGE_MENU:
                self.lang_menu.setTitle(
                    "Language" if lang_norm.startswith("en") else self.tr("Язык")
                )
            if lang_norm.startswith("en"):
                self.act_reset_view.setText("Reset view (defaults)")
                self._act_theme_light.setText("Light")
                self._act_theme_dark.setText("Dark")
                self._act_theme_system.setText("System")
                if SHOW_LANGUAGE_MENU:
                    self._act_lang_ru.setText("Russian")
                    self._act_lang_en.setText("English")
            else:
                self.act_reset_view.setText(self.tr("Сбросить вид (по умолчанию)"))
                self._act_theme_light.setText(self.tr("Светлая"))
                self._act_theme_dark.setText(self.tr("Тёмная"))
                self._act_theme_system.setText(self.tr("Системная"))
                if SHOW_LANGUAGE_MENU:
                    self._act_lang_ru.setText(self.tr("Русский"))
                    self._act_lang_en.setText(self.tr("English"))
        except Exception:
            pass
        try:
            self.help_menu.setTitle(
                "Help" if lang_norm.startswith("en") else self.tr("Помощь")
            )
        except Exception:
            pass
        # Пункты Help
        try:
            if lang_norm.startswith("en"):
                self.act_help.setText("Help")
                self.act_logs.setText("Logs")
                self.act_license.setText("License Agreement")
                self.act_about.setText("About")
            else:
                self.act_help.setText(self.tr("Справка"))
                self.act_logs.setText(self.tr("Логи"))
                self.act_license.setText(self.tr("Лицензионное соглашение"))
                self.act_about.setText(self.tr("О программе"))
        except Exception:
            pass

    def _on_theme_selected(self, theme: str) -> None:
        try:
            self.apply_theme(theme)
            settings = QSettings()
            settings.setValue("ui/theme", theme)
        except Exception:
            pass

    def apply_theme(self, theme: str) -> None:
        app_inst = QApplication.instance()
        if app_inst is None:
            return
        app = cast(QApplication, app_inst)
        theme = (theme or "system").lower()
        # Базовый стиль для жирных заголовков групп
        base_group_qss = "QGroupBox { font-weight: 700; }"
        if theme == "dark":
            dark_qss = """
                QWidget { background-color: #121212; color: #e0e0e0; }
                QGroupBox { border: 1px solid #333; margin-top: 8px; }
                QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top left; padding: 0 3px; }
                QLineEdit, QComboBox, QSpinBox, QDoubleSpinBox, QTextEdit, QPlainTextEdit { background-color: #1e1e1e; color: #e0e0e0; border: 1px solid #444; selection-background-color: #3d6ea1; selection-color: #ffffff; }
                QTableView { background-color: #1e1e1e; gridline-color: #333; alternate-background-color: #151515; }
                QHeaderView::section { background-color: #2a2a2a; color: #e0e0e0; border: 1px solid #444; }
                QPushButton { background-color: #2a2a2a; color: #e0e0e0; border: 1px solid #444; padding: 4px 8px; }
                QPushButton:hover { background-color: #333333; }
                QPushButton:pressed { background-color: #3a3a3a; }
                QMenuBar { background-color: #1e1e1e; }
                QMenuBar::item:selected { background-color: #333333; }
                QMenu { background-color: #1e1e1e; color: #e0e0e0; }
                QMenu::item:selected { background-color: #333333; }
                QCheckBox, QRadioButton { background: transparent; }
                """
            app.setStyleSheet(dark_qss + base_group_qss)
            self._current_theme = "dark"
        elif theme == "light":
            # Светлая тема: системная палитра + наш базовый стиль
            app.setStyleSheet(base_group_qss)
            self._current_theme = "light"
        else:
            # Системная: сбрасываем кастомный стиль кроме жирных заголовков
            app.setStyleSheet(base_group_qss)
            self._current_theme = "system"

    def _on_language_selected(self, lang: str) -> None:
        try:
            lang = (lang or "ru").lower()
            settings = QSettings()
            settings.setValue("ui/language", lang)
            try:
                settings.sync()
            except Exception:
                pass
            # Установим/снимем переводчик на лету
            translator_applied = False
            try:
                translator_applied = self._apply_qtranslator_runtime(lang)
            except Exception:
                translator_applied = False
            # Мгновенно обновим заголовки главного меню
            try:
                self._apply_menu_language(lang)
            except Exception:
                pass
            # Обновим отображение списков компонентов под выбранный язык
            try:
                self.cold_mix.refresh_component_list()
                self.hot_mix.refresh_component_list()
            except Exception:
                pass
            # Обновим заголовок окна и панели
            try:
                if lang.startswith("en"):
                    self.setWindowTitle("Two-stream heat exchanger")
                else:
                    self.setWindowTitle(self.tr("Двухпоточный теплообмен"))
            except Exception:
                pass
            try:
                if hasattr(self, "cold_panel"):
                    self.cold_panel.retranslate_panel(lang)  # type: ignore[attr-defined]
                if hasattr(self, "hot_panel"):
                    self.hot_panel.retranslate_panel(lang)  # type: ignore[attr-defined]
            except Exception:
                pass
            try:
                if hasattr(self, "hydro"):
                    self.hydro.retranslate_panel(lang)  # type: ignore[attr-defined]
            except Exception:
                pass
            try:
                if hasattr(self, "out_panel"):
                    self.out_panel.retranslate_panel(lang)  # type: ignore[attr-defined]
            except Exception:
                pass
            # Обновим панели смесей
            try:
                if hasattr(self, "cold_mix"):
                    self.cold_mix.retranslate_panel(lang)  # type: ignore[attr-defined]
                if hasattr(self, "hot_mix"):
                    self.hot_mix.retranslate_panel(lang)  # type: ignore[attr-defined]
            except Exception:
                pass
            # Обновим заголовки таблиц.
            try:
                self.cold_mix.model.retranslate_headers()
                self.hot_mix.model.retranslate_headers()
            except Exception:
                pass
            # Переименуем существующие строки смесей (RU<->EN)
            try:
                if hasattr(self, "cold_mix"):
                    self.cold_mix.retranslate_existing_rows(lang)  # type: ignore[attr-defined]
                if hasattr(self, "hot_mix"):
                    self.hot_mix.retranslate_existing_rows(lang)  # type: ignore[attr-defined]
            except Exception:
                pass
            # Обновим статус-бар (короткое сообщение)
            try:
                if lang.startswith("en"):
                    self.status.showMessage("Ready")
                else:
                    self.status.showMessage(self.tr("Готово"))
            except Exception:
                pass
            # Кнопки справа (основные действия)
            try:
                if lang.startswith("en"):
                    self.calc_btn.setText("Calculate")
                    self.reset_btn.setText("Clear parameters")
                    self.analysis_btn.setText("Run analysis")
                    self.analysis_btn.setToolTip(
                        "Run analysis by varying component shares"
                    )
                    self.recalc_btn.setText("Recalculate")
                    self.recalc_btn.setToolTip("Recalculate after changes")
                else:
                    self.calc_btn.setText(self.tr("Вычислить"))
                    self.reset_btn.setText(self.tr("Очистить параметры"))
                    self.analysis_btn.setText(self.tr("Провести анализ"))
                    self.analysis_btn.setToolTip(
                        self.tr("Провести анализ изменяя доли компонентов потоков")
                    )
                    self.recalc_btn.setText(self.tr("Перерасчёт"))
                    self.recalc_btn.setToolTip(self.tr("Пересчитать после изменений"))
            except Exception:
                pass
            # Если .qm отсутствует — предупредим пользователя (без перезапуска)
            if not translator_applied and lang.startswith("en"):
                try:
                    base_dir = os.path.dirname(os.path.abspath(__file__))
                    qm_path = os.path.join(base_dir, "i18n", f"HeatSim_{lang}.qm")
                    QMessageBox.information(
                        self,
                        self.tr("Смена языка"),
                        self.tr(
                            "Файл перевода не найден: {p}\nЧасть интерфейса переключена на английский, полная локализация станет доступна после добавления .qm."
                        ).format(p=qm_path),
                    )
                except Exception:
                    pass
        except Exception:
            pass

    def _on_mix_changed(self, *args: Any) -> None:
        """Handler for mix model changes: update calc button and attempt minimal auto-calc."""
        try:
            self._update_calc_button_state()
        except Exception:
            pass
        try:
            self._auto_calc_minimal()
        except Exception:
            pass
        # Любое изменение смеси после первого явного вычисления делает результаты устаревшими
        try:
            if getattr(self, "_explicit_calc_done", False):
                self._results_stale = True
                if self._mix_valid(self.cold_mix.mix_rows()) and self._mix_valid(
                    self.hot_mix.mix_rows()
                ):
                    self.recalc_btn.show()
                    self.calc_btn.hide()
        except Exception:
            pass

    def _remove_initial_focus(self) -> None:
        """Убирает фокус с поля температуры на входе холодного потока при старте."""
        try:
            self.cold_panel.t_in.clearFocus()
            # Переводим фокус на главное окно (ничто не редактируется по умолчанию)
            self.setFocus(Qt.OtherFocusReason)
        except Exception:
            pass

    def _apply_qtranslator_runtime(self, lang: str) -> bool:
        """Устанавливает или снимает переводчик на лету. Возвращает True, если переводчик применён."""
        app_inst = QApplication.instance()
        if app_inst is None:
            return False
        app = cast(QApplication, app_inst)
        # Снимем предыдущий, если был
        try:
            prev = getattr(app, "_app_translator", None)
            if prev is not None:
                app.removeTranslator(prev)
        except Exception:
            pass
        lang = (lang or "ru").lower()
        if not lang.startswith("en"):
            try:
                setattr(app, "_app_translator", None)
                setattr(app, "_app_translator_lang", "")
            except Exception:
                pass
            return True
        try:
            qm_path = str(resource_path("i18n", f"HeatSim_{lang}.qm"))
            if not os.path.exists(qm_path):
                return False
            tr = QTranslator()
            if not tr.load(qm_path):
                return False
            app.installTranslator(tr)
            setattr(app, "_app_translator", tr)
            setattr(app, "_app_translator_lang", str(lang))
            return True
        except Exception:
            return False

    def _on_calc_button_clicked(self) -> None:
        """Handler for explicit user click on Calculate: clear import suppression and run full calculation."""
        try:
            self._suppress_full_calc_after_import = False
        except Exception:
            pass
        success = False
        try:
            success = bool(self.on_calc())
        except Exception:
            success = False
        try:
            if success:
                self._explicit_calc_done = True
                self._results_stale = False
                self.recalc_btn.hide()
                self.calc_btn.show()
                # Save snapshot of inputs after explicit successful calculation
                try:
                    self._last_calc_snapshot = self._relevant_inputs_snapshot()
                except Exception:
                    pass
                try:
                    self.status.showMessage(
                        self.tr("Вычисления выполнены успешно"), 4000
                    )
                except Exception:
                    pass
        except Exception:
            pass

    def _on_schema_changed(self, checked: bool) -> None:
        """Called when hydro schema radio buttons toggle.
        При смене схемы помечаем результаты как устаревшие и показываем кнопку «Перерасчёт».
        """
        try:
            if not checked:
                return
            # Любая смена схемы делает результат устаревшим; не выполняем автопересчёт
            self._results_stale = True
            # показываем кнопку «Перерасчёт» всегда, чтобы пользователь мог подтвердить пересчёт
            try:
                self.recalc_btn.show()
                self.calc_btn.hide()
            except Exception:
                pass
            # Снять подавление «после импорта», чтобы пересчёт по кнопке прошёл нормально
            try:
                self._suppress_full_calc_after_import = False
            except Exception:
                pass
        except Exception:
            pass

    def _gather_inputs_for_export(self) -> Dict[str, Any]:
        """Collect current inputs into a JSON-serializable dict."""
        try:
            cold = self.cold_panel.to_dict()
        except Exception:
            cold = {}
        try:
            hot = self.hot_panel.to_dict()
        except Exception:
            hot = {}
        try:
            cold_mix = self.cold_mix.mix_rows()
        except Exception:
            cold_mix = []
        try:
            hot_mix = self.hot_mix.mix_rows()
        except Exception:
            hot_mix = []
        schema = None
        try:
            schema = self.hydro.current_schema()
        except Exception:
            schema = None
        q_txt = None
        try:
            q_txt = self.out_panel.q.text().strip()
        except Exception:
            q_txt = None
        return {
            "cold": cold,
            "hot": hot,
            "cold_mix": cold_mix,
            "hot_mix": hot_mix,
            "schema": schema,
            "q": q_txt,
        }

    def export_inputs(self) -> None:
        """Export inputs to a CSV file that can be opened in Excel.
        Format: a single CSV with a leading `section` column. Sections:
        - flow_cold: columns t_in,t_out,m,p
        - flow_hot: columns t_in,t_out,m,p
        - mix_cold / mix_hot: columns name,share,tb,cf,cp,rf
        """
        try:
            path, _ = QFileDialog.getSaveFileName(
                self,
                self.tr("Экспорт входных данных"),
                "",
                "CSV Files (*.csv);;All Files (*)",
            )
            if not path:
                return
            data = self._gather_inputs_for_export()

            # Helper to protect Excel from auto-formatting values (dates/numbers)
            def protect_for_excel(val: Any) -> str:
                s = str(val)
                if not s:
                    return ""
                # Common date-like patterns with / or - or mm/dd or dd-mm etc.
                if any(ch in s for ch in ("/", "-")) and any(c.isdigit() for c in s):
                    return "'" + s
                # Leading zeros or long numeric strings -> keep as text
                if s.startswith("0") and len(s) > 1 and s[1].isdigit():
                    return "'" + s
                return s

            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=";")
                # flows
                w.writerow(["section", "t_in", "t_out", "m", "p"])
                w.writerow(["type", "K", "K", "kg/s", "kg/m^2"])
                cold = cast(Dict[str, Any], data.get("cold") or {})
                hot = cast(Dict[str, Any], data.get("hot") or {})
                w.writerow(
                    [
                        "flow_cold",
                        protect_for_excel(cold.get("t_in", "")),
                        protect_for_excel(cold.get("t_out", "")),
                        protect_for_excel(cold.get("m", "")),
                        protect_for_excel(cold.get("p", "")),
                    ]
                )
                w.writerow(
                    [
                        "flow_hot",
                        protect_for_excel(hot.get("t_in", "")),
                        protect_for_excel(hot.get("t_out", "")),
                        protect_for_excel(hot.get("m", "")),
                        protect_for_excel(hot.get("p", "")),
                    ]
                )
                # mixes
                w.writerow([])
                w.writerow(["section", "name", "share", "tb", "cf", "cp", "rf"])
                w.writerow(
                    ["type", "str", "fraction", "K", "kJ/kg*K", "kJ/kg*K", "kJ/kg"]
                )

                def _nz(v: Any) -> str:
                    try:
                        if v in (None, ""):
                            return ""
                        if isinstance(v, (int, float)) and float(v) == 0.0:
                            return ""
                        s = str(v)
                        if s.replace(".", "", 1).isdigit() and float(s) == 0.0:
                            return ""
                        return protect_for_excel(v)
                    except Exception:
                        return ""

                for r in data.get("cold_mix", []):
                    w.writerow(
                        [
                            "mix_cold",
                            protect_for_excel(r.get("name", "")),
                            _nz(r.get("share", "")),
                            _nz(r.get("tb", "")),
                            _nz(r.get("cf", "")),
                            _nz(r.get("cp", "")),
                            _nz(r.get("rf", "")),
                        ]
                    )
                for r in data.get("hot_mix", []):
                    w.writerow(
                        [
                            "mix_hot",
                            protect_for_excel(r.get("name", "")),
                            _nz(r.get("share", "")),
                            _nz(r.get("tb", "")),
                            _nz(r.get("cf", "")),
                            _nz(r.get("cp", "")),
                            _nz(r.get("rf", "")),
                        ]
                    )
                # schema and q
                w.writerow([])
                w.writerow(["meta", "schema", data.get("schema", "")])
                w.writerow(["meta", "q", data.get("q", "")])
        except Exception as e:
            QMessageBox.warning(self, self.tr("Ошибка экспорта"), str(e))

    def export_inputs_xlsx(self) -> None:
        """Export inputs to an .xlsx workbook with separate sheets for flows and mixes."""
        if openpyxl is None:
            QMessageBox.warning(
                self,
                self.tr("Excel экспорт"),
                self.tr("Требуется пакет openpyxl. Установите его в окружение."),
            )
            return
        try:
            path, _ = QFileDialog.getSaveFileName(
                self,
                self.tr("Экспорт в Excel"),
                "",
                "Excel Files (*.xlsx);;All Files (*)",
            )
            if not path:
                return
            data = self._gather_inputs_for_export()
            # create workbook via module to satisfy static checkers
            wb = openpyxl.Workbook()
            # Flows sheet
            ws1 = cast(Any, wb.active)
            ws1.title = "flows"
            ws1.append(["section", "t_in", "t_out", "m", "p"])
            cold = cast(Dict[str, Any], data.get("cold") or {})
            hot = cast(Dict[str, Any], data.get("hot") or {})
            ws1.append(
                [
                    "flow_cold",
                    cold.get("t_in", ""),
                    cold.get("t_out", ""),
                    cold.get("m", ""),
                    cold.get("p", ""),
                ]
            )
            ws1.append(
                [
                    "flow_hot",
                    hot.get("t_in", ""),
                    hot.get("t_out", ""),
                    hot.get("m", ""),
                    hot.get("p", ""),
                ]
            )
            # mixes
            ws2 = wb.create_sheet("mix_cold")
            ws2.append(["name", "share", "tb", "cf", "cp", "rf"])
            for r in data.get("cold_mix", []):
                ws2.append(
                    [
                        r.get("name", ""),
                        r.get("share", ""),
                        r.get("tb", ""),
                        r.get("cf", ""),
                        r.get("cp", ""),
                        r.get("rf", ""),
                    ]
                )
            ws3 = wb.create_sheet("mix_hot")
            ws3.append(["name", "share", "tb", "cf", "cp", "rf"])
            for r in data.get("hot_mix", []):
                ws3.append(
                    [
                        r.get("name", ""),
                        r.get("share", ""),
                        r.get("tb", ""),
                        r.get("cf", ""),
                        r.get("cp", ""),
                        r.get("rf", ""),
                    ]
                )
            # meta
            ws_meta = wb.create_sheet("meta")
            ws_meta.append(["schema", data.get("schema", "")])
            ws_meta.append(["q", data.get("q", "")])
            wb.save(path)
        except Exception as e:
            QMessageBox.warning(self, self.tr("Ошибка экспорта Excel"), str(e))

    def import_inputs_xlsx(self) -> None:
        """Import inputs from an .xlsx workbook created by `export_inputs_xlsx`."""
        if openpyxl is None:
            QMessageBox.warning(
                self,
                self.tr("Excel импорт"),
                self.tr("Требуется пакет openpyxl. Установите его в окружение."),
            )
            return
        try:
            self._importing = True
            path, _ = QFileDialog.getOpenFileName(
                self,
                self.tr("Импорт из Excel"),
                "",
                "Excel Files (*.xlsx);;All Files (*)",
            )
            if not path:
                return
            wb = openpyxl.load_workbook(path, data_only=True)
            assert openpyxl is not None
            # flows
            if "flows" in wb.sheetnames:
                ws = cast(Any, wb["flows"])
                rows = list(ws.values)
                # expect header then two rows
                if len(rows) >= 3:
                    fc = rows[1]
                    fh = rows[2]
                    # safe extraction from tuples that may have variable length
                    from typing import Optional, Sequence

                    def safe_get(
                        cell_tuple: Optional[Sequence[Any]], idx: int, default: str = ""
                    ) -> str:
                        try:
                            if not cell_tuple or len(cell_tuple) <= idx:
                                return default
                            v = cell_tuple[idx]
                            return "" if v is None else str(v)
                        except Exception:
                            return default

                    self.cold_panel.t_in.setText(safe_get(fc, 1))
                    self.cold_panel.t_out.setText(safe_get(fc, 2))
                    self.cold_panel.m.setText(safe_get(fc, 3))
                    self.cold_panel.p.setText(safe_get(fc, 4))

                    self.hot_panel.t_in.setText(safe_get(fh, 1))
                    self.hot_panel.t_out.setText(safe_get(fh, 2))
                    self.hot_panel.m.setText(safe_get(fh, 3))
                    self.hot_panel.p.setText(safe_get(fh, 4))
            # mixes
            for name, target in (
                ("mix_cold", self.cold_mix),
                ("mix_hot", self.hot_mix),
            ):
                if name in wb.sheetnames:
                    try:
                        ws = cast(Any, wb[name])
                        rows = list(ws.values)[1:]
                        if target.model.rowCount() > 0:
                            target.model.removeRows(0, target.model.rowCount())

                        def cell_to_float(val: Any) -> float:
                            try:
                                if val is None:
                                    return 0.0
                                return float(val)
                            except Exception:
                                return 0.0

                        def cell_to_str(val: Any) -> str:
                            try:
                                if val is None:
                                    return ""
                                return str(val)
                            except Exception:
                                return ""

                        import re as _re

                        _date_res = [
                            _re.compile(r"^\d{4}[-/]\d{2}[-/]\d{2}$"),
                            _re.compile(r"^\d{2}[-./]\d{2}[-./]\d{4}$"),
                        ]

                        def _is_bad_share(raw: Any, numeric: float) -> bool:
                            try:
                                if numeric == 0.0:
                                    return True
                                if raw is None:
                                    return False
                                s = str(raw).strip()
                                if _re.search(r"[A-Za-zА-Яа-я]", s):
                                    return True
                                for rg in _date_res:
                                    if rg.match(s):
                                        return True
                                return False
                            except Exception:
                                return True

                        # Определим, нужно ли показывать имена на EN
                        def _to_display_name(name: str) -> str:
                            try:
                                app_inst = QApplication.instance()
                                active_lang = (
                                    str(
                                        getattr(app_inst, "_app_translator_lang", "")
                                        or ""
                                    ).lower()
                                    if app_inst is not None
                                    else ""
                                )
                                desired = str(
                                    QSettings().value("ui/language", "ru") or "ru"
                                ).lower()
                                to_en = desired == "en" or active_lang.startswith("en")
                            except Exception:
                                to_en = False
                            # Определяем русский ключ компонента
                            ru_key = name
                            try:
                                if name in COMPONENT_DB:
                                    ru_key = name
                                elif name in COMPONENT_NAME_RU_FROM_EN:
                                    ru_key = COMPONENT_NAME_RU_FROM_EN[name]
                            except Exception:
                                pass
                            return (
                                COMPONENT_NAME_EN.get(ru_key, ru_key)
                                if to_en
                                else ru_key
                            )

                        for row in rows:
                            try:
                                nm = cell_to_str(row[0]) if row and len(row) > 0 else ""
                                raw_share = row[1] if row and len(row) > 1 else None
                                share = cell_to_float(raw_share)
                                if _is_bad_share(raw_share, share):
                                    continue
                                tb = (
                                    cell_to_float(row[2])
                                    if row and len(row) > 2
                                    else 0.0
                                )
                                cf = (
                                    cell_to_float(row[3])
                                    if row and len(row) > 3
                                    else 0.0
                                )
                                cp = (
                                    cell_to_float(row[4])
                                    if row and len(row) > 4
                                    else 0.0
                                )
                                rf = (
                                    cell_to_float(row[5])
                                    if row and len(row) > 5
                                    else 0.0
                                )
                                display_nm = _to_display_name(nm)
                                target.model.add_or_update(
                                    display_nm, share, tb, cf, cp, rf
                                )
                            except Exception:
                                pass
                    except Exception:
                        pass
            # meta
            if "meta" in wb.sheetnames:
                try:
                    ws = wb["meta"]
                    rows = list(ws.values)
                    for r in rows:
                        if not r or len(r) == 0:
                            continue
                        if len(r) > 0 and r[0] == "schema":
                            schema_val = r[1] if len(r) > 1 else None
                            if schema_val:
                                if schema_val == "Schema1":
                                    self.hydro.rb_mix_mix.setChecked(True)
                                elif schema_val == "Schema2":
                                    self.hydro.rb_parallel.setChecked(True)
                                elif schema_val == "Schema3":
                                    self.hydro.rb_mix_cold.setChecked(True)
                                elif schema_val == "Schema4":
                                    self.hydro.rb_mix_hot.setChecked(True)
                                elif schema_val == "Schema5":
                                    self.hydro.rb_counter.setChecked(True)
                        if len(r) > 0 and r[0] == "q":
                            try:
                                if len(r) > 1:
                                    self.out_panel.q.setText(str(r[1] or ""))
                            except Exception:
                                pass
                except Exception:
                    pass
            try:
                self._try_auto_calc()
                self._update_calc_button_state()
            except Exception:
                pass
            # Перевести заголовки и имена компонентов под активный язык
            try:
                self.cold_mix.model.retranslate_headers()
                self.hot_mix.model.retranslate_headers()
                lang_now = str(QSettings().value("ui/language", "ru") or "ru").lower()
                self.cold_mix.retranslate_existing_rows(lang_now)
                self.hot_mix.retranslate_existing_rows(lang_now)
            except Exception:
                pass
            try:
                self._suppress_full_calc_after_import = True
            except Exception:
                pass
            try:
                self._post_import_changed = False
                self.recalc_btn.hide()
                self._lock_imported_fields()
            except Exception:
                pass
        except Exception as e:
            QMessageBox.warning(self, self.tr("Ошибка импорта Excel"), str(e))
        finally:
            try:
                self._importing = False
            except Exception:
                pass

    def import_inputs(self) -> None:
        """Import inputs from a CSV file with the format written by export_inputs().
        During import only minimal auto-fill is performed (t_out or q). Full sigma/K
        calculation is not executed; user should press "Вычислить" to compute σ и K.
        """
        import re

        date_regexes = [
            re.compile(r"^\d{4}[-/]\d{2}[-/]\d{2}$"),  # YYYY-MM-DD or YYYY/MM/DD
            re.compile(
                r"^\d{2}[-./]\d{2}[-./]\d{4}$"
            ),  # DD-MM-YYYY / DD.MM.YYYY / DD/MM/YYYY
        ]

        def _is_invalid_token(tok: str) -> bool:
            t = tok.strip()
            if not t:
                return False
            # letters present
            if re.search(r"[A-Za-zА-Яа-я]", t):
                return True
            # date like patterns
            for rg in date_regexes:
                if rg.match(t):
                    return True
            return False

        flows: Dict[str, Dict[str, str]] = {"flow_cold": {}, "flow_hot": {}}
        cold_mix: List[Dict[str, Any]] = []
        hot_mix: List[Dict[str, Any]] = []
        schema_val: Optional[str] = None
        q_val: Optional[str] = None
        try:
            self._importing = True
            path, _ = QFileDialog.getOpenFileName(
                self,
                self.tr("Импорт входных данных"),
                "",
                "CSV Files (*.csv);;All Files (*)",
            )
            if not path:
                return
            with open(path, "r", encoding="utf-8-sig") as f:
                rdr = csv.reader(f, delimiter=";")
                for row in rdr:
                    if not row:
                        continue
                    first = row[0].strip()
                    if first.lower() == "section":
                        continue
                    if first in ("flow_cold", "flow_hot"):
                        # strip leading apostrophe inserted to protect Excel-format
                        def strip_protect(x: str) -> str:
                            if x.startswith("'"):
                                return x[1:]
                            return x

                        val_t_in = strip_protect(row[1]) if len(row) > 1 else ""
                        val_t_out = strip_protect(row[2]) if len(row) > 2 else ""
                        val_m = strip_protect(row[3]) if len(row) > 3 else ""
                        val_p = strip_protect(row[4]) if len(row) > 4 else ""
                        flows[first]["t_in"] = (
                            "" if _is_invalid_token(val_t_in) else val_t_in
                        )
                        flows[first]["t_out"] = (
                            "" if _is_invalid_token(val_t_out) else val_t_out
                        )
                        flows[first]["m"] = "" if _is_invalid_token(val_m) else val_m
                        flows[first]["p"] = "" if _is_invalid_token(val_p) else val_p
                        continue
                    if first in ("mix_cold", "mix_hot"):
                        name = row[1] if len(row) > 1 else ""
                        share = row[2] if len(row) > 2 else "0"
                        tb = row[3] if len(row) > 3 else "0"
                        cf = row[4] if len(row) > 4 else "0"
                        cp = row[5] if len(row) > 5 else "0"
                        rf = row[6] if len(row) > 6 else "0"
                        # если числовые поля не валидны (слово/дата) – делаем их пустыми чтобы fallback -> 0.0
                        for var_name, raw in [
                            ("share", share),
                            ("tb", tb),
                            ("cf", cf),
                            ("cp", cp),
                            ("rf", rf),
                        ]:
                            if _is_invalid_token(raw):
                                if var_name == "share":
                                    share = ""
                                elif var_name == "tb":
                                    tb = ""
                                elif var_name == "cf":
                                    cf = ""
                                elif var_name == "cp":
                                    cp = ""
                                elif var_name == "rf":
                                    rf = ""
                        # проверка доли: если не число или 0 -> пропускаем компонент
                        share_is_valid = False
                        try:
                            share_val_tmp = (
                                float(share.replace(",", ".")) if share.strip() else 0.0
                            )
                            if share_val_tmp != 0.0:
                                share_is_valid = True
                        except Exception:
                            share_is_valid = False
                        if not share_is_valid:
                            continue
                        try:
                            share_f = float(share.replace(",", "."))
                            tb_f = float(tb.replace(",", "."))
                            cf_f = float(cf.replace(",", "."))
                            cp_f = float(cp.replace(",", "."))
                            rf_f = float(rf.replace(",", "."))
                            rec = {
                                "name": name,
                                "share": share_f,
                                "tb": tb_f,
                                "cf": cf_f,
                                "cp": cp_f,
                                "rf": rf_f,
                            }
                        except Exception:
                            # любое исключение при парсинге -> полностью пропускаем компонент
                            continue
                        if first == "mix_cold":
                            cold_mix.append(rec)
                        else:
                            hot_mix.append(rec)
                        continue
                    if first == "meta":
                        key = row[1] if len(row) > 1 else ""
                        val = row[2] if len(row) > 2 else ""
                        if key == "schema":
                            schema_val = val
                        elif key == "q":
                            q_val = val

            # populate flows
            fc = flows.get("flow_cold", {})
            if fc:
                self.cold_panel.t_in.setText(str(fc.get("t_in", "")))
                self.cold_panel.t_out.setText(str(fc.get("t_out", "")))
                self.cold_panel.m.setText(str(fc.get("m", "")))
                self.cold_panel.p.setText(str(fc.get("p", "")))
            fh = flows.get("flow_hot", {})
            if fh:
                self.hot_panel.t_in.setText(str(fh.get("t_in", "")))
                self.hot_panel.t_out.setText(str(fh.get("t_out", "")))
                self.hot_panel.m.setText(str(fh.get("m", "")))
                self.hot_panel.p.setText(str(fh.get("p", "")))

            # replace mixes (с учётом активного языка для отображения имён)
            if self.cold_mix.model.rowCount() > 0:
                self.cold_mix.model.removeRows(0, self.cold_mix.model.rowCount())

            def _to_display_name(name: str) -> str:
                try:
                    app_inst = QApplication.instance()
                    active_lang = (
                        str(getattr(app_inst, "_app_translator_lang", "") or "").lower()
                        if app_inst is not None
                        else ""
                    )
                    desired = str(
                        QSettings().value("ui/language", "ru") or "ru"
                    ).lower()
                    to_en = desired == "en" or active_lang.startswith("en")
                except Exception:
                    to_en = False
                ru_key = name
                try:
                    if name in COMPONENT_DB:
                        ru_key = name
                    elif name in COMPONENT_NAME_RU_FROM_EN:
                        ru_key = COMPONENT_NAME_RU_FROM_EN[name]
                except Exception:
                    pass
                return COMPONENT_NAME_EN.get(ru_key, ru_key) if to_en else ru_key

            for r in cold_mix:
                try:
                    self.cold_mix.model.add_or_update(
                        _to_display_name(str(r.get("name", ""))),
                        float(r.get("share", 0.0) or 0.0),
                        float(r.get("tb", 0.0) or 0.0),
                        float(r.get("cf", 0.0) or 0.0),
                        float(r.get("cp", 0.0) or 0.0),
                        float(r.get("rf", 0.0) or 0.0),
                    )
                except Exception:
                    pass
            if self.hot_mix.model.rowCount() > 0:
                self.hot_mix.model.removeRows(0, self.hot_mix.model.rowCount())
            for r in hot_mix:
                try:
                    self.hot_mix.model.add_or_update(
                        _to_display_name(str(r.get("name", ""))),
                        float(r.get("share", 0.0) or 0.0),
                        float(r.get("tb", 0.0) or 0.0),
                        float(r.get("cf", 0.0) or 0.0),
                        float(r.get("cp", 0.0) or 0.0),
                        float(r.get("rf", 0.0) or 0.0),
                    )
                except Exception:
                    pass

            # set schema and q if present
            if schema_val:
                if schema_val == "Schema1":
                    self.hydro.rb_mix_mix.setChecked(True)
                elif schema_val == "Schema2":
                    self.hydro.rb_parallel.setChecked(True)
                elif schema_val == "Schema3":
                    self.hydro.rb_mix_cold.setChecked(True)
                elif schema_val == "Schema4":
                    self.hydro.rb_mix_hot.setChecked(True)
                elif schema_val == "Schema5":
                    self.hydro.rb_counter.setChecked(True)
            if q_val is not None:
                self.out_panel.q.setText(str(q_val))

            # normalize and attempt minimal auto-calc
            try:
                self._try_auto_calc()
                self._update_calc_button_state()
            except Exception:
                pass
            # Перевести заголовки и имена компонентов под активный язык
            try:
                self.cold_mix.model.retranslate_headers()
                self.hot_mix.model.retranslate_headers()
                lang_now = str(QSettings().value("ui/language", "ru") or "ru").lower()
                self.cold_mix.retranslate_existing_rows(lang_now)
                self.hot_mix.retranslate_existing_rows(lang_now)
            except Exception:
                pass
            # suppress full sigma/K calculation on subsequent schema toggles until user confirms
            try:
                self._suppress_full_calc_after_import = True
            except Exception:
                pass
            try:
                self._post_import_changed = False
                self.recalc_btn.hide()
                self._lock_imported_fields()
            except Exception:
                pass
        except Exception as e:
            QMessageBox.warning(self, self.tr("Ошибка импорта"), str(e))
        finally:
            try:
                self._importing = False
            except Exception:
                pass

    # ===================== ИМПОРТ/ЭКСПОРТ БАЗЫ КОМПОНЕНТОВ =====================
    def import_component_db_csv(self) -> None:
        try:
            path, _ = QFileDialog.getOpenFileName(
                self,
                self.tr("Импорт базы компонентов (CSV)"),
                str(DATA_DIR),
                "CSV Files (*.csv);;All Files (*)",
            )
            if not path:
                return
            stats = load_component_db_from_csv(path, merge=True)
            # обновить выпадающие списки компонентов в обеих панелях смесей
            try:
                self.cold_mix.refresh_component_list()
            except Exception:
                pass
            try:
                self.hot_mix.refresh_component_list()
            except Exception:
                pass
            QMessageBox.information(
                self,
                self.tr("Импорт базы компонентов"),
                self.tr(
                    "Добавлено: {added}\nОбновлено: {updated}\nПропущено: {skipped}"
                ).format(
                    added=stats.get("added", 0),
                    updated=stats.get("updated", 0),
                    skipped=stats.get("skipped", 0),
                ),
            )
        except Exception as e:
            QMessageBox.warning(
                self, self.tr("Ошибка импорта базы компонентов"), str(e)
            )

    def import_component_db_xlsx(self) -> None:
        if openpyxl is None:
            QMessageBox.warning(
                self,
                self.tr("Excel импорт"),
                self.tr("Для импорта из Excel требуется пакет openpyxl."),
            )
            return
        try:
            path, _ = QFileDialog.getOpenFileName(
                self,
                self.tr("Импорт базы компонентов (Excel)"),
                str(DATA_DIR),
                "Excel Files (*.xlsx);;All Files (*)",
            )
            if not path:
                return
            stats = load_component_db_from_xlsx(path, merge=True)
            try:
                self.cold_mix.refresh_component_list()
            except Exception:
                pass
            try:
                self.hot_mix.refresh_component_list()
            except Exception:
                pass
            QMessageBox.information(
                self,
                self.tr("Импорт базы компонентов"),
                self.tr(
                    "Добавлено: {added}\nОбновлено: {updated}\nПропущено: {skipped}"
                ).format(
                    added=stats.get("added", 0),
                    updated=stats.get("updated", 0),
                    skipped=stats.get("skipped", 0),
                ),
            )
        except Exception as e:
            QMessageBox.warning(
                self, self.tr("Ошибка импорта базы компонентов"), str(e)
            )

    # CSV-экспорт базы компонентов удалён; используйте экспорт в Excel.

    def export_component_db_xlsx(self) -> None:
        if openpyxl is None:
            QMessageBox.warning(
                self,
                self.tr("Excel экспорт"),
                self.tr("Для экспорта в Excel требуется пакет openpyxl."),
            )
            return
        try:
            path, _ = QFileDialog.getSaveFileName(
                self,
                self.tr("Экспорт базы компонентов (Excel)"),
                str(DATA_DIR / "components.xlsx"),
                "Excel Files (*.xlsx);;All Files (*)",
            )
            if not path:
                return
            export_component_db_to_xlsx(path)
            QMessageBox.information(
                self, self.tr("Экспорт базы компонентов"), self.tr("Готово.")
            )
        except Exception as e:
            QMessageBox.warning(
                self, self.tr("Ошибка экспорта базы компонентов"), str(e)
            )

    def _can_compute_sigma_k(self) -> bool:
        """Return True if we have enough validated inputs to compute sigma and k."""
        try:
            cold = self.cold_panel.to_dict()
            hot = self.hot_panel.to_dict()
            cold_mix = self.cold_mix.mix_rows()
            hot_mix = self.hot_mix.mix_rows()
            # require either Q or hot T_out present (the calculate() can derive missing t_out from Q)
            q_present = bool(self.out_panel.q.text().strip())
            hot_tout_present = bool(self.hot_panel.t_out.text().strip())
            if not (q_present or hot_tout_present):
                return False

            # require cold stream to have t_in, t_out, m and valid mix
            cold_ok = (
                bool(cold.get("t_in"))
                and bool(cold.get("t_out"))
                and bool(cold.get("m"))
                and self._mix_valid(cold_mix)
            )

            # For hot stream: if hot_tout is present, require t_in, t_out and m; if only Q is present,
            # require at least t_in and m (t_out may be computed by calculate()).
            hot_has_tin = bool(hot.get("t_in"))
            hot_has_m = bool(hot.get("m"))
            if hot_tout_present:
                hot_ok = (
                    hot_has_tin
                    and bool(hot.get("t_out"))
                    and hot_has_m
                    and self._mix_valid(hot_mix)
                )
            else:
                hot_ok = hot_has_tin and hot_has_m and self._mix_valid(hot_mix)

            return cold_ok and hot_ok
        except Exception:
            return False

    @staticmethod
    def _mix_valid(mix: Sequence[MixRow]) -> bool:
        try:
            if not mix:
                return False
            s = sum(float(item.get("share", 0.0)) for item in mix)
            return abs(s - 1.0) <= 1e-3
        except Exception:
            return False

    def _update_calc_button_state(self) -> None:
        """Highlight `self.calc_btn` when sigma/k can be computed by pressing it."""
        try:
            ready = self._can_compute_sigma_k()
            if ready:
                # highlight: yellow background and bold
                self.calc_btn.setStyleSheet("background: #ffec8b; font-weight: 700;")
            else:
                self.calc_btn.setStyleSheet("")
        except Exception:
            pass

    # --- блокировка полей ---
    def _on_q_changed(self) -> None:
        has_q = self.out_panel.q.text().strip() != ""
        # when Q has value, disable T_out (hot)
        set_enabled(self.hot_panel.t_out, not has_q)

    def _on_tplus_out_changed(self) -> None:
        has_tout = self.hot_panel.t_out.text().strip() != ""
        # when T_out has value, disable Q
        set_enabled(self.out_panel.q, not has_tout)

    def _on_q_edit_finished(self) -> None:
        # при завершении ввода Q — блокируем T+out и попытка вычислить T_out автоматически
        try:
            has_q = self.out_panel.q.text().strip() != ""
            set_enabled(self.hot_panel.t_out, not has_q)
            if has_q:
                # вызовем calculate и если вернётся t_out_plus — заполним
                cold = self.cold_panel.to_dict()
                hot = self.hot_panel.to_dict()
                cold_mix = self.cold_mix.mix_rows()
                hot_mix = self.hot_mix.mix_rows()
                q_val = to_float(self.out_panel.q.text())
                res = getattr(logic, "calculate", None)
                if callable(res):
                    ans = res(
                        cold=cold,
                        hot=hot,
                        cold_mix=cold_mix,
                        hot_mix=hot_mix,
                        q=q_val,
                        schema=self.hydro.current_schema(),
                    )
                    ans = cast(Dict[str, Any], ans)
                    ans_dict = ans
                    if ans_dict and "t_out_plus" in ans_dict:
                        # временно блокируем сигналы при записи
                        self.hot_panel.t_out.blockSignals(True)
                        self.hot_panel.t_out.setText(f"{ans_dict['t_out_plus']:.6g}")
                        self.hot_panel.t_out.blockSignals(False)
                        try:
                            self._update_calc_button_state()
                        except Exception:
                            pass
        except Exception:
            pass

    def _on_tplus_out_edit_finished(self) -> None:
        # при завершении ввода T+out — блокируем Q и попытка вычислить Q автоматически
        try:
            has_tout = self.hot_panel.t_out.text().strip() != ""
            set_enabled(self.out_panel.q, not has_tout)
            if has_tout:
                cold = self.cold_panel.to_dict()
                hot = self.hot_panel.to_dict()
                cold_mix = self.cold_mix.mix_rows()
                hot_mix = self.hot_mix.mix_rows()
                q_val = to_float(self.out_panel.q.text())
                res = getattr(logic, "calculate", None)
                if callable(res):
                    ans = res(
                        cold=cold,
                        hot=hot,
                        cold_mix=cold_mix,
                        hot_mix=hot_mix,
                        q=q_val,
                        schema=self.hydro.current_schema(),
                    )
                    ans = cast(Dict[str, Any], ans)
                    ans_dict = ans
                    if ans_dict and "q" in ans_dict:
                        self.out_panel.q.blockSignals(True)
                        self.out_panel.q.setText(f"{ans_dict['q']:.6g}")
                        self.out_panel.q.blockSignals(False)
                        try:
                            self._update_calc_button_state()
                        except Exception:
                            pass
        except Exception:
            pass

    def on_calc(self) -> bool:
        # Программные вызовы не блокируем — пересчёт должен выполняться при необходимости.
        cold = self.cold_panel.to_dict()
        hot = self.hot_panel.to_dict()
        cold_mix = self.cold_mix.mix_rows()
        hot_mix = self.hot_mix.mix_rows()
        # Раньше требовалось обязательное Q или T_out(hot). Теперь доверим logic.calculate
        # попытаться вывести недостающие величины, если это возможно.
        # Дополнительно: если ни Q, ни T⁺out не указаны — не запускаем расчёт и показываем подсказку,
        # чтобы не создавать впечатление успешного пустого расчёта.
        try:
            q_present = bool(self.out_panel.q.text().strip())
            hot_tout_present = bool(self.hot_panel.t_out.text().strip())
            if not (q_present or hot_tout_present):
                try:
                    app_inst = QApplication.instance()
                    active_lang = (
                        str(getattr(app_inst, "_app_translator_lang", "") or "").lower()
                        if app_inst
                        else ""
                    )
                    desired = str(
                        QSettings().value("ui/language", "ru") or "ru"
                    ).lower()
                    is_en = active_lang.startswith("en") or desired == "en"
                except Exception:
                    is_en = False
                if is_en:
                    title = "Calculation"
                    text = (
                        "Enter either heat load Q or hot stream outlet temperature T⁺out — "
                        "one of them is required."
                    )
                else:
                    title = self.tr("Расчёт")
                    text = self.tr(
                        "Введите тепловую нагрузку Q или температуру выхода горячего потока T⁺out — "
                        "требуется минимум одно из этих значений."
                    )
                QMessageBox.information(self, title, text)
                return False
        except Exception:
            pass
        # примечание: ранее использовалась локальная переменная q_text для проверок,
        # сейчас логика рассчитывает недостающие значения без предварительного отказа
        q_val = to_float(self.out_panel.q.text())
        schema = self.hydro.current_schema()

        try:
            res = getattr(logic, "calculate", None)
            if callable(res):
                ans = res(
                    cold=cold,
                    hot=hot,
                    cold_mix=cold_mix,
                    hot_mix=hot_mix,
                    q=q_val,
                    schema=schema,
                )
                ans = cast(Dict[str, Any], ans)
                if ans:
                    # safely extract numeric/string values from ans
                    try:
                        sigma_val = (
                            float(ans.get("sigma", 0.0))
                            if ans.get("sigma") is not None
                            else 0.0
                        )
                    except Exception:
                        sigma_val = 0.0
                    try:
                        k_val = (
                            float(ans.get("k", 0.0))
                            if ans.get("k") is not None
                            else 0.0
                        )
                    except Exception:
                        k_val = 0.0
                    # Обновляем значения всегда, даже если они равны 0.0, чтобы не оставались старые значения
                    self.out_panel.sigma.setText(format_num(sigma_val))
                    set_enabled(self.out_panel.sigma, False)
                    self.out_panel.k.setText(format_num(k_val))
                    set_enabled(self.out_panel.k, False)

                    # статус (не использовать слово "Schema" в выводе)
                    schema_display = schema
                    contact = ""
                    q_show = q_val
                    s_show = 0.0
                    try:
                        k_src = str(ans.get("k_source", ""))
                        contact = str(ans.get("contact_type", ""))
                        q_show = float(ans.get("q", q_val) or q_val)
                        k_show = k_val
                        s_show = sigma_val
                        schema_display = str(ans.get("schema", schema))
                        msg = f"{schema_display}  contact={contact or '-'}  k_source={k_src or '-'}  Q={q_show:.4g}  K={k_show:.4g}  σ={s_show:.4g}"
                        self.status.showMessage(msg)
                    except Exception:
                        pass
                    set_enabled(self.out_panel.k, False)
                    # Если σ посчитана, но K отсутствует, попробуем ещё раз выполнить расчёт (возможно
                    # теперь доступны дополнительные данные после записи t_out или q) и получить K.
                    try:
                        if ans and ("sigma" in ans) and (not ans.get("k")):
                            # reload inputs (t_out/q might have been filled above)
                            cold2 = self.cold_panel.to_dict()
                            hot2 = self.hot_panel.to_dict()
                            cold_mix2 = self.cold_mix.mix_rows()
                            hot_mix2 = self.hot_mix.mix_rows()
                            q2 = float(ans.get("q", q_val) or q_val)
                            res2 = getattr(logic, "calculate", None)
                            if callable(res2):
                                ans2 = res2(
                                    cold=cold2,
                                    hot=hot2,
                                    cold_mix=cold_mix2,
                                    hot_mix=hot_mix2,
                                    q=q2,
                                    schema=schema,
                                )
                                ans2 = cast(Dict[str, Any], ans2)
                                if ans2 and ("k" in ans2):
                                    try:
                                        k2_val = float(ans2.get("k", 0.0) or 0.0)
                                    except Exception:
                                        k2_val = 0.0
                                    self.out_panel.k.setText(format_num(k2_val))
                                    set_enabled(self.out_panel.k, False)
                                    # update status with new k info
                                    try:
                                        k_src2 = str(ans2.get("k_source", ""))
                                        k_show2 = k2_val
                                        msg2 = f"{schema_display}  contact={contact or '-'}  k_source={k_src2 or '-'}  Q={q_show:.4g}  K={k_show2:.4g}  σ={s_show:.4g}"
                                        self.status.showMessage(msg2)
                                    except Exception:
                                        pass
                    except Exception:
                        pass
                if ans and "q" in ans:
                    # записываем Q и делаем T+out недоступным для ввода
                    try:
                        self.out_panel.q.blockSignals(True)
                        self.out_panel.q.setText(f"{ans['q']:.6g}")
                        self.out_panel.q.blockSignals(False)
                    except Exception:
                        pass
                    set_enabled(self.hot_panel.t_out, False)
                if ans and "t_out_plus" in ans:
                    self.hot_panel.t_out.setText(f"{ans['t_out_plus']:.6g}")
                    set_enabled(self.out_panel.q, False)
                # persist schema selection (store current hydro schema id)
                try:
                    schema_path = os.path.join(
                        os.path.dirname(os.path.abspath(__file__)),
                        "data",
                        "csv",
                        "schema.txt",
                    )
                    with open(schema_path, "w", encoding="utf-8") as f:
                        f.write(self.hydro.current_schema())
                except Exception:
                    pass
            else:
                QMessageBox.information(
                    self,
                    self.tr("Расчёт"),
                    self.tr("Функция расчёта в logic.py не найдена. Заполните её."),
                )
                return False
        except Exception as e:
            QMessageBox.warning(self, self.tr("Ошибка расчёта"), str(e))
            return False
        return True

    def _auto_calc_minimal(self) -> None:
        """Выполнить быстрый авторасчёт, только для заполнения `q` или `t_out_plus`.
        Не обновляет `sigma` и `k` — эти величины вычисляются только по нажатию кнопки.
        """
        cold = self.cold_panel.to_dict()
        hot = self.hot_panel.to_dict()
        cold_mix = self.cold_mix.mix_rows()
        hot_mix = self.hot_mix.mix_rows()
        q_val = to_float(self.out_panel.q.text())
        schema = self.hydro.current_schema()
        try:
            res = getattr(logic, "calculate", None)
            if callable(res):
                ans = res(
                    cold=cold,
                    hot=hot,
                    cold_mix=cold_mix,
                    hot_mix=hot_mix,
                    q=q_val,
                    schema=schema,
                )
                # Only apply q or t_out_plus if provided by calculation
                ans = cast(Dict[str, Any], ans)
                ans_dict = ans
                if (
                    ans_dict
                    and "q" in ans_dict
                    and (not self.out_panel.q.text().strip())
                ):
                    try:
                        self.out_panel.q.blockSignals(True)
                        self.out_panel.q.setText(f"{ans['q']:.6g}")
                        self.out_panel.q.blockSignals(False)
                    except Exception:
                        pass
                if (
                    ans_dict
                    and "t_out_plus" in ans_dict
                    and (not self.hot_panel.t_out.text().strip())
                ):
                    try:
                        self.hot_panel.t_out.blockSignals(True)
                        self.hot_panel.t_out.setText(f"{ans_dict['t_out_plus']:.6g}")
                        self.hot_panel.t_out.blockSignals(False)
                    except Exception:
                        pass
                # update button state after potential auto-fill
                try:
                    self._update_calc_button_state()
                except Exception:
                    pass
        except Exception:
            pass

    # --- сброс данных ---
    def on_reset(self) -> None:
        """Полный сброс входных параметров и результатов в UI.
        Очищает поля потоков, таблицы смесей и поля результатов (Q, σ, K).
        """
        try:
            # clear flow panels
            for w in (
                self.cold_panel.t_in,
                self.cold_panel.t_out,
                self.cold_panel.m,
                self.cold_panel.p,
                self.hot_panel.t_in,
                self.hot_panel.t_out,
                self.hot_panel.m,
                self.hot_panel.p,
            ):
                try:
                    w.blockSignals(True)
                    w.clear()
                    w.blockSignals(False)
                    set_enabled(w, True)
                except Exception:
                    pass
            try:
                self._post_import_changed = False
                self._suppress_full_calc_after_import = False
                self.recalc_btn.hide()
                self.calc_btn.show()
            except Exception:
                pass

            # clear mixtures
            try:
                if self.cold_mix.model.rowCount() > 0:
                    self.cold_mix.model.removeRows(0, self.cold_mix.model.rowCount())
                if self.hot_mix.model.rowCount() > 0:
                    self.hot_mix.model.removeRows(0, self.hot_mix.model.rowCount())
                # сброс выбора компонента к первому (обычно "Азот")
                try:
                    self.cold_mix.comp.setCurrentIndex(0)
                except Exception:
                    pass
                try:
                    self.hot_mix.comp.setCurrentIndex(0)
                except Exception:
                    pass
                # очистка поля ввода доли
                try:
                    self.cold_mix.share.clear()
                except Exception:
                    pass
                try:
                    self.hot_mix.share.clear()
                except Exception:
                    pass
                # сброс отображения суммы
                try:
                    self.cold_mix.sum_field.setText("0.0")
                except Exception:
                    pass
                try:
                    self.hot_mix.sum_field.setText("0.0")
                except Exception:
                    pass
                # update hints/export
                try:
                    self.cold_mix.update_share_hint()
                except Exception:
                    pass
                try:
                    self.hot_mix.update_share_hint()
                except Exception:
                    pass
            except Exception:
                pass

            # clear outputs
            try:
                self.out_panel.clear_values()
                try:
                    set_enabled(self.out_panel.q, True)
                except Exception:
                    pass
            except Exception:
                # fallback: direct resets
                try:
                    self.out_panel.q.clear()
                    try:
                        set_enabled(self.out_panel.q, True)
                    except Exception:
                        pass
                except Exception:
                    pass
                try:
                    self.out_panel.sigma.setText("0.0")
                    set_enabled(self.out_panel.sigma, False)
                except Exception:
                    pass
                try:
                    self.out_panel.k.setText("0.0")
                    set_enabled(self.out_panel.k, False)
                except Exception:
                    pass

            # reset status
            try:
                self.status.showMessage(self.tr("Сброшено"))
            except Exception:
                pass
            # Сброс схемы к первой и запрет автосчёта sigma/k до явного вычисления
            try:
                self.hydro.rb_mix_mix.setChecked(True)
            except Exception:
                pass
            try:
                self._explicit_calc_done = False
            except Exception:
                pass
        except Exception:
            pass

    # --- Диалоги помощи ---
    def _simple_text_dialog(
        self, title: str, text: str, read_only: bool = True, force_en: bool = False
    ) -> None:
        try:
            dlg = QDialog(self)
            dlg.setWindowTitle(title)
            dlg.resize(700, 500)
            layout = QVBoxLayout(dlg)
            te = QTextEdit()
            te.setPlainText(text)
            te.setReadOnly(read_only)
            layout.addWidget(te)
            buttons = QDialogButtonBox(QDialogButtonBox.Close)
            # Подписываем кнопку закрытия в зависимости от языка
            try:
                if force_en:
                    buttons.button(QDialogButtonBox.Close).setText("Close")
                else:
                    buttons.button(QDialogButtonBox.Close).setText(self.tr("Закрыть"))
            except Exception:
                pass
            buttons.rejected.connect(dlg.reject)
            buttons.accepted.connect(dlg.accept)
            layout.addWidget(buttons)
            dlg.exec_()
        except Exception as e:
            QMessageBox.warning(self, title, str(e))

    def show_help_dialog(self) -> None:
        # Определяем, нужно ли принудительно показывать англ. текст (если выбран EN, но переводчик не загружен)
        force_en = False
        try:
            app_inst = QApplication.instance()
            active_lang = (
                str(getattr(app_inst, "_app_translator_lang", "") or "").lower()
                if app_inst is not None
                else ""
            )
            settings = QSettings()
            desired = str(settings.value("ui/language", "ru") or "ru").lower()
            force_en = (desired == "en") and (not active_lang.startswith("en"))
        except Exception:
            force_en = False

        if force_en:
            help_text = (
                "User guide:\n\n"
                "1. Enter parameters for cold and hot streams (temperatures, flow rate, pressure).\n"
                "2. Create mixtures: select a component, enter its share, and add. The sum of shares for each mixture must be 1.\n"
                "3. Select the hydrodynamic scheme.\n"
                "4. Enter either heat load Q or hot stream outlet temperature T⁺out — the other will be calculated automatically.\n"
                "5. Press 'Calculate' to get σ and K.\n"
                "6. The 'Run analysis' button opens a separate window for changing shares and plotting the Q–σ curve.\n"
                "7. Use the 'File' menu to import/export data in CSV or Excel. Imported values are not converted to dates.\n"
                "8. 'Clear parameters' resets all fields.\n"
            )
            self._simple_text_dialog("User guide", help_text, force_en=True)
        else:
            help_text = (
                self.tr("Справка по использованию программы:")
                + "\n\n"
                + self.tr(
                    "1. Введите параметры холодного и горячего потоков (температуры, расход, давление)."
                )
                + "\n"
                + self.tr(
                    "2. Сформируйте смеси компонентов: выберите компонент, долю и добавьте. Сумма долей каждой смеси должна быть 1."
                )
                + "\n"
                + self.tr("3. Выберите гидродинамическую схему.")
                + "\n"
                + self.tr(
                    "4. Введите либо тепловую нагрузку Q, либо выходную температуру горячего потока T⁺out — второе значение будет рассчитано автоматически."
                )
                + "\n"
                + self.tr("5. Нажмите 'Вычислить' для получения σ и K.")
                + "\n"
                + self.tr(
                    "6. Кнопка 'Провести анализ' позволяет открыть отдельное окно для изменения долей и построения графика зависимости Q–σ."
                )
                + "\n"
                + self.tr(
                    "7. Используйте меню 'Файл' для импорта/экспорта данных в CSV или Excel. При импорте значения не преобразуются в даты."
                )
                + "\n"
                + self.tr("8. 'Очистить параметры' сбрасывает все поля.")
                + "\n"
            )
            self._simple_text_dialog(self.tr("Справка"), help_text)

    def show_logs_dialog(self) -> None:
        try:
            if LOG_FILE.exists():
                content = LOG_FILE.read_text(encoding="utf-8", errors="ignore")
            else:
                content = self.tr("Логи отсутствуют.")
            self._simple_text_dialog(self.tr("Логи"), content)
        except Exception as e:
            QMessageBox.warning(self, self.tr("Логи"), str(e))

    def show_license_dialog(self) -> None:
        try:
            lic_path = resource_path("Лицензионное_соглашение.txt")
            if not lic_path.exists():
                QMessageBox.information(
                    self,
                    self.tr("Лицензионное соглашение"),
                    self.tr("Файл лицензионного соглашения не найден."),
                )
                return
            try:
                content = lic_path.read_text(encoding="utf-8")
            except Exception as e:
                QMessageBox.warning(
                    self,
                    self.tr("Лицензионное соглашение"),
                    self.tr("Не удалось прочитать файл: {err}").format(err=e),
                )
                return
            self._simple_text_dialog(self.tr("Лицензионное соглашение"), content)
        except Exception as e:
            QMessageBox.warning(self, self.tr("Лицензионное соглашение"), str(e))

    def show_about_dialog(self) -> None:
        try:
            # Определяем режим англ. текста при отсутствии переводчика
            force_en = False
            try:
                app_inst = QApplication.instance()
                active_lang = (
                    str(getattr(app_inst, "_app_translator_lang", "") or "").lower()
                    if app_inst is not None
                    else ""
                )
                settings = QSettings()
                desired = str(settings.value("ui/language", "ru") or "ru").lower()
                force_en = (desired == "en") and (not active_lang.startswith("en"))
            except Exception:
                force_en = False

            version_path = resource_path("VERSION")
            version = "неизвестно"
            if version_path.exists():
                try:
                    version = version_path.read_text(encoding="utf-8").strip()
                except Exception:
                    pass
            try:
                if getattr(sys, "frozen", False):
                    # In PyInstaller onefile use executable timestamp
                    mtime_src = Path(sys.executable)
                else:
                    mtime_src = Path(__file__)
                mtime = datetime.fromtimestamp(mtime_src.stat().st_mtime).strftime(
                    "%Y-%m-%d %H:%M"
                )
            except Exception:
                mtime = "unknown"

            if force_en:
                text = (
                    f"Full name: Two-stream heat exchanger analysis program\n"
                    f"Version: {version}\n"
                    f"Last update: {mtime}\n\n"
                    "Description: Tool for calculating heat load,\n"
                    "entropy production, and heat transfer coefficient\n"
                    "in heat exchange systems with various hydrodynamic schemes."
                )
                QMessageBox.information(self, "About", text)
            else:
                text = (
                    f"Полное наименование: Программа анализа двухпоточного теплообменника\n"
                    f"Версия: {version}\n"
                    f"Дата обновления: {mtime}\n\n"
                    "Описание: Инструмент для расчёта тепловой нагрузки,\n"
                    "производства энтропии и коэффициента теплопередачи\n"
                    "в системах теплообмена с различными гидродинамическими схемами."
                )
                QMessageBox.information(self, self.tr("О программе"), text)
        except Exception as e:
            QMessageBox.warning(self, self.tr("О программе"), str(e))

    # --- Окно анализа ---
    def open_analysis_window(self) -> None:
        # Не открывать окно анализа, если выходные параметры неизвестны или устарели
        try:
            explicit_done = bool(getattr(self, "_explicit_calc_done", False))
            results_stale = bool(getattr(self, "_results_stale", False))
            # язык для сообщений
            try:
                app_inst = QApplication.instance()
                active_lang = (
                    str(getattr(app_inst, "_app_translator_lang", "") or "").lower()
                    if app_inst
                    else ""
                )
                desired = str(QSettings().value("ui/language", "ru") or "ru").lower()
                is_en = active_lang.startswith("en") or desired == "en"
            except Exception:
                is_en = False
            if not explicit_done:
                if is_en:
                    QMessageBox.information(
                        self,
                        "Analysis",
                        "Run 'Calculate' first — output parameters (Q or T⁺out, σ and K) are required for analysis.",
                    )
                else:
                    QMessageBox.information(
                        self,
                        self.tr("Анализ"),
                        self.tr(
                            "Сначала выполните расчёт — выходные параметры (Q или T⁺out, σ и K) необходимы для анализа."
                        ),
                    )
                return
            if results_stale:
                if is_en:
                    QMessageBox.information(
                        self,
                        "Analysis",
                        "Inputs were changed. Press 'Recalculate' before opening analysis.",
                    )
                else:
                    QMessageBox.information(
                        self,
                        self.tr("Анализ"),
                        self.tr(
                            "Входные данные изменены. Нажмите 'Перерасчёт' перед открытием анализа."
                        ),
                    )
                return
        except Exception:
            # Если проверка не удалась, не блокируем, но стараемся продолжить с существующей логикой
            pass
        try:
            from analysis_interface import AnalysisWindow  # type: ignore
        except Exception as e:
            QMessageBox.warning(
                self,
                self.tr("Анализ"),
                self.tr("Не удалось импортировать окно анализа: {e}").format(e=e),
            )
            return
        try:
            cold = self.cold_panel.to_dict()
            hot = self.hot_panel.to_dict()
            cold_mix = self.cold_mix.mix_rows()
            hot_mix = self.hot_mix.mix_rows()
        except Exception as e:
            QMessageBox.warning(
                self,
                self.tr("Анализ"),
                self.tr("Ошибка чтения входных данных: {e}").format(e=e),
            )
            return
        try:
            # приведение типов для mypy/pyright
            from typing import cast as _cast, Dict as _Dict, Any as _Any, List as _List

            cold_cast = {str(k): float(v) for k, v in cold.items()}  # type: ignore[arg-type]
            hot_cast = {str(k): float(v) for k, v in hot.items()}  # type: ignore[arg-type]
            cold_list = [dict(r) for r in cold_mix]  # type: ignore[list-item]
            hot_list = [dict(r) for r in hot_mix]  # type: ignore[list-item]
            cold_list = _cast(_List[_Dict[str, _Any]], cold_list)
            hot_list = _cast(_List[_Dict[str, _Any]], hot_list)
            self._analysis_win = AnalysisWindow(
                cold_flow=cold_cast,
                hot_flow=hot_cast,
                cold_mix=cold_list,
                hot_mix=hot_list,
                schema=self.hydro.current_schema(),
                parent=self,
            )
            self._analysis_win.show()
        except Exception as e:
            QMessageBox.warning(
                self,
                self.tr("Анализ"),
                self.tr("Ошибка открытия окна анализа: {e}").format(e=e),
            )

    # ---------- ПОМЕТКА УСТАРЕВАНИЯ РЕЗУЛЬТАТОВ ----------
    def _mark_stale_results(self) -> None:
        """Пометить, что результаты (σ, K, производные расчёты) устарели после изменения входных данных.
        Кнопка 'Перерасчёт' отображается только если уже был выполнен явный расчёт.
        """
        try:
            if getattr(self, "_explicit_calc_done", False):
                # If user just unlocked a widget but didn't type, ignore this event
                try:
                    unlock_widgets = (
                        self.cold_panel.t_in,
                        self.cold_panel.t_out,
                        self.cold_panel.m,
                        self.hot_panel.t_in,
                        self.hot_panel.t_out,
                        self.hot_panel.m,
                        self.out_panel.q,
                    )
                    for w in unlock_widgets:
                        if getattr(w, "_just_unlocked_waiting", False) and not getattr(
                            w, "_just_unlocked_typed", False
                        ):
                            return
                except Exception:
                    pass

                # Only mark as stale if inputs actually differ from last calculated snapshot
                self._results_stale = True
                try:
                    # compute current snapshot and compare
                    cur = self._relevant_inputs_snapshot()
                    last = getattr(self, "_last_calc_snapshot", None)
                    if last is None:
                        # no previous successful calculation snapshot -> cannot determine delta
                        # всё равно дадим возможность пересчёта
                        self.recalc_btn.show()
                        self.calc_btn.hide()
                        return
                    else:
                        if (
                            cur != last
                            and self._mix_valid(self.cold_mix.mix_rows())
                            and self._mix_valid(self.hot_mix.mix_rows())
                        ):
                            # Показать кнопку пересчёта — пересчёт только по нажатию
                            self.recalc_btn.show()
                            self.calc_btn.hide()
                except Exception:
                    # fallback to old behaviour on any error
                    if self._mix_valid(self.cold_mix.mix_rows()) and self._mix_valid(
                        self.hot_mix.mix_rows()
                    ):
                        self.recalc_btn.show()
                        self.calc_btn.hide()
        except Exception:
            pass

    def _on_recalc_clicked(self) -> None:
        # Ensure suppression flags are cleared before explicit recalculation
        try:
            self._suppress_full_calc_after_import = False
            self._post_import_changed = False
            self._results_stale = False
        except Exception:
            pass

        # Перед явным пересчётом попробуем минимально автозаполнить недостающие Q/T+out
        try:
            self._auto_calc_minimal()
        except Exception:
            pass

        success = False
        res_val = None
        err_msg = None
        try:
            # Выполняем полный расчёт (как при кнопке Вычислить) and capture result/exception
            try:
                res_val = self.on_calc()
            except Exception as e:
                res_val = None
                err_msg = str(e)
            success = bool(res_val)
        except Exception:
            success = False

        try:
            if success:
                # on success, hide the recalc button and save snapshot
                try:
                    self._last_calc_snapshot = self._relevant_inputs_snapshot()
                except Exception:
                    pass
                try:
                    self.recalc_btn.hide()
                    self.calc_btn.show()
                except Exception:
                    pass
                try:
                    self._explicit_calc_done = True
                except Exception:
                    pass
            else:
                # keep the button visible and inform the user with diagnostics
                try:
                    # gather some useful internal state for diagnostics
                    try:
                        q_txt = self.out_panel.q.text().strip()
                    except Exception:
                        q_txt = "<error>"
                    try:
                        t_out_txt = self.hot_panel.t_out.text().strip()
                    except Exception:
                        t_out_txt = "<error>"
                    try:
                        can_compute = bool(self._can_compute_sigma_k())
                    except Exception:
                        can_compute = False
                    try:
                        cur_snap = self._relevant_inputs_snapshot()
                    except Exception:
                        cur_snap = {}
                    last_snap = getattr(self, "_last_calc_snapshot", None)
                    diag = (
                        self.tr("Перерасчёт не выполнен.")
                        + "\n\n"
                        + self.tr("Внутреннее состояние:")
                        + "\n"
                        + f"_suppress_full_calc_after_import={getattr(self, '_suppress_full_calc_after_import', None)}\n"
                        + f"_post_import_changed={getattr(self, '_post_import_changed', None)}\n"
                        + f"_results_stale={getattr(self, '_results_stale', None)}\n"
                        + f"Q='{q_txt}'  T_out='{t_out_txt}'  can_compute={can_compute}\n\n"
                        + f"last_snapshot={last_snap}\n"
                        + f"current_snapshot={cur_snap}\n\n"
                        + f"on_calc returned: {res_val!r}\n"
                        + f"exception: {err_msg or '<none>'}"
                    )
                    QMessageBox.warning(self, self.tr("Перерасчёт"), diag)
                except Exception:
                    # fallback simple message
                    try:
                        QMessageBox.warning(
                            self,
                            self.tr("Перерасчёт"),
                            self.tr(
                                "Перерасчёт не выполнен — проверьте входные данные или сообщение об ошибке."
                            ),
                        )
                    except Exception:
                        pass
        except Exception:
            pass

    def _relevant_inputs_snapshot(self) -> Dict[str, Any]:
        """Return a small dict snapshot of input values that affect σ/K calculations.
        Used to determine whether inputs changed since last explicit calculation.
        """
        try:
            snap: Dict[str, Any] = {}

            # numeric fields that affect sigma/K (ignore pressure 'p' which is not used)
            def num_of(widget: QLineEdit) -> Optional[float]:
                try:
                    return round(float(widget.text().strip().replace(",", ".")), 6)
                except Exception:
                    return None

            snap["cold_t_in"] = num_of(self.cold_panel.t_in)
            snap["cold_t_out"] = num_of(self.cold_panel.t_out)
            snap["cold_m"] = num_of(self.cold_panel.m)
            snap["hot_t_in"] = num_of(self.hot_panel.t_in)
            snap["hot_t_out"] = num_of(self.hot_panel.t_out)
            snap["hot_m"] = num_of(self.hot_panel.m)
            snap["q"] = num_of(self.out_panel.q)
            # include simple representation of mixes (shares and cf/cp) to detect significant mix changes
            try:

                from typing import Mapping as _Mapping

                def norm_row(r: _Mapping[str, Any]) -> Dict[str, Optional[float]]:
                    try:
                        return {
                            "share": round(float(r.get("share", 0.0)), 6),
                            "cf": round(float(r.get("cf", 0.0)), 6),
                            "cp": round(float(r.get("cp", 0.0)), 6),
                        }
                    except Exception:
                        return {"share": None, "cf": None, "cp": None}

                cm = [norm_row(r) for r in self.cold_mix.mix_rows()]
                hm = [norm_row(r) for r in self.hot_mix.mix_rows()]
            except Exception:
                cm = []
                hm = []
            snap["cold_mix"] = cm
            snap["hot_mix"] = hm
            snap["schema"] = self.hydro.current_schema()
            return snap
        except Exception:
            return {}

    def _lock_imported_fields(self) -> None:
        """Заблокировать поля, заполненные при импорте. Разблокировка при очистке."""
        try:
            widgets = [
                self.cold_panel.t_in,
                self.cold_panel.t_out,
                self.cold_panel.m,
                self.cold_panel.p,
                self.hot_panel.t_in,
                self.hot_panel.t_out,
                self.hot_panel.m,
                self.hot_panel.p,
                self.out_panel.q,
            ]
            for w in widgets:
                if w.text().strip():
                    set_enabled(w, False)
        except Exception:
            pass

    def _normalize_input(self) -> None:
        """Нормализация числовых полей: замена запятой, удаление лишних ведущих нулей (кроме '0.'), очистка одиночного '0'."""
        edits = [
            self.cold_panel.t_in,
            self.cold_panel.t_out,
            self.cold_panel.m,
            self.hot_panel.t_in,
            self.hot_panel.t_out,
            self.hot_panel.m,
            self.out_panel.q,
        ]
        for w in edits:
            try:
                txt = w.text().strip()
                if not txt:
                    continue
                txt = txt.replace(",", ".")
                # если формат вроде 09 или 00012.3 -> убираем ведущие нули
                if txt.count(".") <= 1:
                    if (
                        txt.startswith("0")
                        and len(txt) > 1
                        and not txt.startswith("0.")
                    ):
                        # убрать все ведущие нули, оставить один перед точкой если была
                        if "." in txt:
                            int_part, frac_part = txt.split(".", 1)
                            int_part = int_part.lstrip("0") or "0"
                            txt = int_part + (
                                "." + frac_part if frac_part != "" else ""
                            )
                        else:
                            txt = txt.lstrip("0") or "0"
                w.blockSignals(True)
                w.setText(txt)
                w.blockSignals(False)
            except Exception:
                pass
        # any manual edit should lift the import-based suppression of full calculation
        try:
            self._suppress_full_calc_after_import = False
        except Exception:
            pass
        # после нормализации ничего не считаем автоматически
        try:
            self._update_calc_button_state()
        except Exception:
            pass

    def _try_auto_calc(self) -> None:
        """Попытаться выполнить расчёт автоматически (вызывается после editingFinished важных полей)."""
        try:
            # Проверим, есть ли все необходимые входы для автоматического вычисления
            cold = self.cold_panel.to_dict()
            hot = self.hot_panel.to_dict()
            cold_mix = self.cold_mix.mix_rows()
            hot_mix = self.hot_mix.mix_rows()
            q_text = self.out_panel.q.text().strip()
            t_out_hot_text = self.hot_panel.t_out.text().strip()

            # use shared _mix_valid which accepts Sequence[MixRow]

            # 1) If Q is empty and we have sufficient data in either stream -> compute Q
            if not q_text:
                cold_ready = (
                    cold["t_in"] and cold["t_out"] and cold["m"]
                ) and self._mix_valid(cold_mix)
                hot_ready = (
                    hot["t_in"] and hot["t_out"] and hot["m"]
                ) and self._mix_valid(hot_mix)
                if cold_ready or hot_ready:
                    self._auto_calc_minimal()
                    return

            # 2) If t_out_hot is empty but Q is given and hot stream data + mix are valid -> compute t_out_hot
            if not t_out_hot_text and q_text:
                hot_ready_for_tout = (hot["t_in"] and hot["m"]) and self._mix_valid(
                    hot_mix
                )
                if hot_ready_for_tout:
                    self._auto_calc_minimal()
                    return
            # otherwise do nothing
        except Exception:
            pass

    def _on_any_input_changed(self) -> None:
        """Реакция на любое завершение ввода ключевых полей после успешного явного расчёта.
        Показывает кнопку «Перерасчёт», выполняет авто‑дозаполнение и полный пересчёт значений
        без обновления снимка, оставляя результаты помеченными как устаревшие до подтверждения.
        """
        try:
            if not getattr(self, "_explicit_calc_done", False):
                return
            # показать кнопку «Перерасчёт», скрыть «Вычислить»
            try:
                self._results_stale = True
                self.recalc_btn.show()
                self.calc_btn.hide()
            except Exception:
                pass
            # Только помечаем как устаревшие и ждём явного подтверждения пользователем
            try:
                self._suppress_full_calc_after_import = False
            except Exception:
                pass
        except Exception:
            pass

    def reset_view(self) -> None:
        """Сброс окна к виду по умолчанию: обычное состояние и базовый размер."""
        try:
            # Сброс состояния окна и возврат к базовому размеру
            self.showNormal()
            self.resize(1600, 975)
            # Центрирование на текущем экране
            try:
                scr = self.screen()  # type: ignore[attr-defined]
            except Exception:
                scr = None
            if scr is None:
                try:
                    scr = QApplication.primaryScreen()
                except Exception:
                    scr = None
            if scr is not None:
                try:
                    geo = scr.availableGeometry()
                    x = geo.x() + (geo.width() - self.width()) // 2
                    y = geo.y() + (geo.height() - self.height()) // 2
                    self.move(max(geo.left(), x), max(geo.top(), y))
                except Exception:
                    pass
            self.status.showMessage(self.tr("Вид сброшен к значению по умолчанию"))
        except Exception:
            pass


"""Модуль interface: содержит классы GUI (панели и главное окно).

Точка входа приложения перенесена в main.py.
"""
